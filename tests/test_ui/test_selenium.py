"""Selenium end-to-end tests for the HTMX UI.

Tests run with a visible Chrome browser to demonstrate UI functionality.
Example values are loaded from the MIAPPE YAML spec.
"""

import subprocess
import time

import pytest
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from tests.test_ui.examples import (
    BIO_MAT_EXAMPLE,
    DATA_FILE_EXAMPLE,
    ENVIRONMENT_EXAMPLE,
    EVENT_EXAMPLE,
    FACTOR_EXAMPLE,
    FACTOR_VALUE_EXAMPLE,
    INV_EXAMPLE,
    OBS_UNIT_EXAMPLE,
    OBS_VAR_EXAMPLE,
    PERSON_EXAMPLE,
    SAMPLE_EXAMPLE,
    STUDY_EXAMPLE,
)

# Delay constants per requirements
FILL_DELAY = 0.1  # Delay between form fills
CLICK_DELAY = 0.5  # Delay after button clicks

BASE_URL = "http://127.0.0.1:8081"


@pytest.fixture(scope="module")
def server():
    """Start the MIAPPE-API server for testing."""
    import os
    import socket

    # Set working directory to miappe-api root
    cwd = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    proc = subprocess.Popen(
        ["uv", "run", "uvicorn", "metaseed.ui.routes:app", "--port", "8081"],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        cwd=cwd,
    )

    # Wait for server to be ready by polling the port
    max_attempts = 30
    for _ in range(max_attempts):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex(("127.0.0.1", 8081))
            sock.close()
            if result == 0:
                break
        except Exception:
            pass
        time.sleep(0.5)
    else:
        # If server failed to start, print output for debugging
        output = proc.stdout.read().decode() if proc.stdout else ""
        proc.terminate()
        raise RuntimeError(f"Server failed to start within timeout. Output: {output}")

    time.sleep(0.5)  # Extra buffer for full initialization
    yield proc
    proc.terminate()
    proc.wait()


@pytest.fixture
def browser(server):  # noqa: ARG001
    """Create a visible Chrome browser for testing."""
    _ = server  # Ensure server is running
    options = Options()
    # Run in visible mode (no headless)
    options.add_argument("--window-size=1280,900")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(5)

    # Reset server state before each test
    import urllib.request

    try:
        req = urllib.request.Request(f"{BASE_URL}/reset", method="POST")
        urllib.request.urlopen(req, timeout=5)
    except Exception:
        pass  # Ignore errors if server not ready

    yield driver
    driver.quit()


def fill_field(driver, testid: str, value: str, trigger_change: bool = False):
    """Fill a form field by data-testid.

    Args:
        driver: Selenium WebDriver
        testid: The data-testid attribute value
        value: The value to fill
        trigger_change: If True, trigger a change event after filling (needed for HTMX)
    """
    element = driver.find_element(By.CSS_SELECTOR, f"[data-testid='{testid}']")

    # For date inputs, use JavaScript to set value directly (avoids locale issues)
    if element.get_attribute("type") == "date":
        driver.execute_script("arguments[0].value = arguments[1]", element, value)
    else:
        element.clear()
        element.send_keys(value)

    if trigger_change:
        # Trigger change event for HTMX handlers
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('change', {bubbles: true}))", element
        )
        time.sleep(CLICK_DELAY)  # Wait for HTMX to process

    time.sleep(FILL_DELAY)


def click_button(driver, testid: str):
    """Click a button by data-testid and wait."""
    button = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, f"[data-testid='{testid}']"))
    )
    button.click()
    time.sleep(CLICK_DELAY)


def click_element(driver, testid: str):
    """Click any element by data-testid and wait."""
    element = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, f"[data-testid='{testid}']"))
    )
    element.click()
    time.sleep(CLICK_DELAY)


def element_exists(driver, testid: str) -> bool:
    """Check if an element with given data-testid exists."""
    elements = driver.find_elements(By.CSS_SELECTOR, f"[data-testid='{testid}']")
    return len(elements) > 0


def get_element_text(driver, testid: str) -> str:
    """Get the text content of an element by data-testid."""
    element = driver.find_element(By.CSS_SELECTOR, f"[data-testid='{testid}']")
    return element.text


def start_new_investigation(driver, profile: str = "miappe"):
    """Start creating a new Investigation by clicking button and selecting profile.

    Args:
        driver: Selenium WebDriver
        profile: Profile to select ("miappe", "isa", or "isa-miappe-combined")
    """
    click_button(driver, "btn-new-investigation")
    click_button(driver, f"profile-{profile}")


def fill_all_study_fields(driver, row_idx: int = 0):
    """Fill all Study fields from YAML example in a table row.

    Args:
        driver: Selenium WebDriver
        row_idx: Row index for cell naming (default 0)
    """
    prefix = f"cell-{row_idx}-"

    # Required fields
    fill_field(driver, f"{prefix}unique_id", STUDY_EXAMPLE["unique_id"], trigger_change=True)
    fill_field(driver, f"{prefix}title", STUDY_EXAMPLE["title"], trigger_change=True)

    # All optional fields from YAML example
    fill_field(driver, f"{prefix}description", STUDY_EXAMPLE["description"], trigger_change=True)
    fill_field(driver, f"{prefix}start_date", STUDY_EXAMPLE["start_date"], trigger_change=True)
    fill_field(driver, f"{prefix}end_date", STUDY_EXAMPLE["end_date"], trigger_change=True)
    fill_field(
        driver,
        f"{prefix}contact_institution",
        STUDY_EXAMPLE["contact_institution"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}experimental_site_name",
        STUDY_EXAMPLE["experimental_site_name"],
        trigger_change=True,
    )
    fill_field(driver, f"{prefix}latitude", str(STUDY_EXAMPLE["latitude"]), trigger_change=True)
    fill_field(driver, f"{prefix}longitude", str(STUDY_EXAMPLE["longitude"]), trigger_change=True)
    fill_field(driver, f"{prefix}altitude", str(STUDY_EXAMPLE["altitude"]), trigger_change=True)
    fill_field(
        driver,
        f"{prefix}growth_facility_type",
        STUDY_EXAMPLE["growth_facility_type"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}experimental_design_type",
        STUDY_EXAMPLE["experimental_design_type"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}experimental_design_description",
        STUDY_EXAMPLE["experimental_design_description"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}observation_unit_description",
        STUDY_EXAMPLE["observation_unit_description"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}cultural_practices",
        STUDY_EXAMPLE["cultural_practices"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}map_of_experimental_design",
        STUDY_EXAMPLE["map_of_experimental_design"],
        trigger_change=True,
    )

    # List field: observation_unit_level_hierarchy (one per line in textarea)
    hierarchy = STUDY_EXAMPLE.get("observation_unit_level_hierarchy", [])
    if hierarchy:
        fill_field(
            driver,
            f"{prefix}observation_unit_level_hierarchy",
            "\n".join(hierarchy),
            trigger_change=True,
        )


def fill_all_person_fields(driver, row_idx: int = 0):
    """Fill all Person fields from YAML example in a table row.

    Args:
        driver: Selenium WebDriver
        row_idx: Row index for cell naming (default 0)
    """
    prefix = f"cell-{row_idx}-"

    # Required field
    fill_field(driver, f"{prefix}name", PERSON_EXAMPLE["name"], trigger_change=True)

    # All optional fields from YAML example
    fill_field(driver, f"{prefix}email", PERSON_EXAMPLE["email"], trigger_change=True)
    fill_field(driver, f"{prefix}institution", PERSON_EXAMPLE["institution"], trigger_change=True)
    fill_field(driver, f"{prefix}role", PERSON_EXAMPLE["role"], trigger_change=True)
    fill_field(driver, f"{prefix}orcid", PERSON_EXAMPLE["orcid"], trigger_change=True)


def fill_all_biological_material_fields(driver, row_idx: int = 0):
    """Fill all BiologicalMaterial fields from YAML example in a table row.

    Args:
        driver: Selenium WebDriver
        row_idx: Row index for cell naming (default 0)
    """
    prefix = f"cell-{row_idx}-"

    # Required field
    fill_field(driver, f"{prefix}unique_id", BIO_MAT_EXAMPLE["unique_id"], trigger_change=True)

    # All optional fields from YAML example
    fill_field(driver, f"{prefix}organism", BIO_MAT_EXAMPLE["organism"], trigger_change=True)
    fill_field(driver, f"{prefix}genus", BIO_MAT_EXAMPLE["genus"], trigger_change=True)
    fill_field(driver, f"{prefix}species", BIO_MAT_EXAMPLE["species"], trigger_change=True)
    fill_field(
        driver,
        f"{prefix}infraspecific_name",
        BIO_MAT_EXAMPLE["infraspecific_name"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}accession_number",
        BIO_MAT_EXAMPLE["accession_number"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}biological_material_description",
        BIO_MAT_EXAMPLE["biological_material_description"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}biological_material_latitude",
        str(BIO_MAT_EXAMPLE["biological_material_latitude"]),
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}biological_material_longitude",
        str(BIO_MAT_EXAMPLE["biological_material_longitude"]),
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}biological_material_altitude",
        str(BIO_MAT_EXAMPLE["biological_material_altitude"]),
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}biological_material_coordinates_uncertainty",
        BIO_MAT_EXAMPLE["biological_material_coordinates_uncertainty"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}biological_material_preprocessing",
        BIO_MAT_EXAMPLE["biological_material_preprocessing"],
        trigger_change=True,
    )

    # List field: external_references (one per line in textarea)
    ext_refs = BIO_MAT_EXAMPLE.get("external_references", [])
    if ext_refs:
        fill_field(driver, f"{prefix}external_references", "\n".join(ext_refs), trigger_change=True)


def fill_all_observation_unit_fields(driver, row_idx: int = 0):
    """Fill all ObservationUnit fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required field
    fill_field(driver, f"{prefix}unique_id", OBS_UNIT_EXAMPLE["unique_id"], trigger_change=True)

    # Optional fields
    fill_field(
        driver,
        f"{prefix}observation_unit_type",
        OBS_UNIT_EXAMPLE["observation_unit_type"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}biological_material_id",
        OBS_UNIT_EXAMPLE["biological_material_id"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}spatial_distribution_type",
        OBS_UNIT_EXAMPLE["spatial_distribution_type"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}spatial_distribution",
        OBS_UNIT_EXAMPLE["spatial_distribution"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}observation_unit_x_ref",
        OBS_UNIT_EXAMPLE["observation_unit_x_ref"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}observation_unit_y_ref",
        OBS_UNIT_EXAMPLE["observation_unit_y_ref"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}observation_unit_block",
        OBS_UNIT_EXAMPLE["observation_unit_block"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}observation_unit_replicate",
        OBS_UNIT_EXAMPLE["observation_unit_replicate"],
        trigger_change=True,
    )
    fill_field(driver, f"{prefix}entry_type", OBS_UNIT_EXAMPLE["entry_type"], trigger_change=True)
    fill_field(
        driver,
        f"{prefix}observation_level",
        OBS_UNIT_EXAMPLE["observation_level"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}observation_level_code",
        OBS_UNIT_EXAMPLE["observation_level_code"],
        trigger_change=True,
    )

    # List field
    ext_refs = OBS_UNIT_EXAMPLE.get("external_references", [])
    if ext_refs:
        fill_field(driver, f"{prefix}external_references", "\n".join(ext_refs), trigger_change=True)


def fill_all_observed_variable_fields(driver, row_idx: int = 0):
    """Fill all ObservedVariable fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required fields
    fill_field(driver, f"{prefix}unique_id", OBS_VAR_EXAMPLE["unique_id"], trigger_change=True)
    fill_field(driver, f"{prefix}name", OBS_VAR_EXAMPLE["name"], trigger_change=True)

    # Optional fields
    fill_field(driver, f"{prefix}trait", OBS_VAR_EXAMPLE["trait"], trigger_change=True)
    fill_field(
        driver,
        f"{prefix}trait_accession_number",
        OBS_VAR_EXAMPLE["trait_accession_number"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}trait_description",
        OBS_VAR_EXAMPLE["trait_description"],
        trigger_change=True,
    )
    fill_field(driver, f"{prefix}method", OBS_VAR_EXAMPLE["method"], trigger_change=True)
    fill_field(
        driver,
        f"{prefix}method_accession_number",
        OBS_VAR_EXAMPLE["method_accession_number"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}method_description",
        OBS_VAR_EXAMPLE["method_description"],
        trigger_change=True,
    )
    fill_field(driver, f"{prefix}scale", OBS_VAR_EXAMPLE["scale"], trigger_change=True)
    fill_field(
        driver,
        f"{prefix}scale_accession_number",
        OBS_VAR_EXAMPLE["scale_accession_number"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}scale_description",
        OBS_VAR_EXAMPLE["scale_description"],
        trigger_change=True,
    )
    fill_field(driver, f"{prefix}time_scale", OBS_VAR_EXAMPLE["time_scale"], trigger_change=True)
    fill_field(driver, f"{prefix}reference", OBS_VAR_EXAMPLE["reference"], trigger_change=True)


def fill_all_factor_fields(driver, row_idx: int = 0):
    """Fill all Factor fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required fields
    fill_field(driver, f"{prefix}unique_id", FACTOR_EXAMPLE["unique_id"], trigger_change=True)
    fill_field(driver, f"{prefix}name", FACTOR_EXAMPLE["name"], trigger_change=True)

    # Optional fields
    fill_field(driver, f"{prefix}description", FACTOR_EXAMPLE["description"], trigger_change=True)
    fill_field(driver, f"{prefix}factor_type", FACTOR_EXAMPLE["factor_type"], trigger_change=True)


def fill_all_factor_value_fields(driver, row_idx: int = 0):
    """Fill all FactorValue fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required fields
    fill_field(driver, f"{prefix}unique_id", FACTOR_VALUE_EXAMPLE["unique_id"], trigger_change=True)
    fill_field(driver, f"{prefix}value", FACTOR_VALUE_EXAMPLE["value"], trigger_change=True)

    # Optional fields
    fill_field(driver, f"{prefix}factor_id", FACTOR_VALUE_EXAMPLE["factor_id"], trigger_change=True)
    fill_field(
        driver, f"{prefix}description", FACTOR_VALUE_EXAMPLE["description"], trigger_change=True
    )


def fill_all_event_fields(driver, row_idx: int = 0):
    """Fill all Event fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required fields
    fill_field(driver, f"{prefix}unique_id", EVENT_EXAMPLE["unique_id"], trigger_change=True)
    fill_field(driver, f"{prefix}event_type", EVENT_EXAMPLE["event_type"], trigger_change=True)

    # Optional fields
    fill_field(driver, f"{prefix}date", EVENT_EXAMPLE["date"], trigger_change=True)
    fill_field(driver, f"{prefix}end_date", EVENT_EXAMPLE["end_date"], trigger_change=True)
    fill_field(driver, f"{prefix}description", EVENT_EXAMPLE["description"], trigger_change=True)
    fill_field(
        driver,
        f"{prefix}event_accession_number",
        EVENT_EXAMPLE["event_accession_number"],
        trigger_change=True,
    )

    # List field
    ou_ids = EVENT_EXAMPLE.get("observation_unit_ids", [])
    if ou_ids:
        fill_field(driver, f"{prefix}observation_unit_ids", "\n".join(ou_ids), trigger_change=True)


def fill_all_environment_fields(driver, row_idx: int = 0):
    """Fill all Environment fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required fields
    fill_field(driver, f"{prefix}unique_id", ENVIRONMENT_EXAMPLE["unique_id"], trigger_change=True)
    fill_field(driver, f"{prefix}parameter", ENVIRONMENT_EXAMPLE["parameter"], trigger_change=True)

    # Optional fields
    fill_field(
        driver,
        f"{prefix}parameter_accession_number",
        ENVIRONMENT_EXAMPLE["parameter_accession_number"],
        trigger_change=True,
    )
    fill_field(driver, f"{prefix}value", ENVIRONMENT_EXAMPLE["value"], trigger_change=True)
    fill_field(driver, f"{prefix}unit", ENVIRONMENT_EXAMPLE["unit"], trigger_change=True)
    fill_field(driver, f"{prefix}date", ENVIRONMENT_EXAMPLE["date"], trigger_change=True)
    fill_field(
        driver, f"{prefix}description", ENVIRONMENT_EXAMPLE["description"], trigger_change=True
    )


def fill_all_sample_fields(driver, row_idx: int = 0):
    """Fill all Sample fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required field
    fill_field(driver, f"{prefix}unique_id", SAMPLE_EXAMPLE["unique_id"], trigger_change=True)

    # Optional fields
    fill_field(
        driver,
        f"{prefix}observation_unit_id",
        SAMPLE_EXAMPLE["observation_unit_id"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}plant_structural_development_stage",
        SAMPLE_EXAMPLE["plant_structural_development_stage"],
        trigger_change=True,
    )
    fill_field(
        driver,
        f"{prefix}plant_anatomical_entity",
        SAMPLE_EXAMPLE["plant_anatomical_entity"],
        trigger_change=True,
    )
    fill_field(
        driver, f"{prefix}collection_date", SAMPLE_EXAMPLE["collection_date"], trigger_change=True
    )
    fill_field(driver, f"{prefix}description", SAMPLE_EXAMPLE["description"], trigger_change=True)

    # List field
    ext_refs = SAMPLE_EXAMPLE.get("external_references", [])
    if ext_refs:
        fill_field(driver, f"{prefix}external_references", "\n".join(ext_refs), trigger_change=True)


def fill_all_data_file_fields(driver, row_idx: int = 0):
    """Fill all DataFile fields from YAML example in a table row."""
    prefix = f"cell-{row_idx}-"

    # Required fields
    fill_field(driver, f"{prefix}unique_id", DATA_FILE_EXAMPLE["unique_id"], trigger_change=True)
    fill_field(driver, f"{prefix}name", DATA_FILE_EXAMPLE["name"], trigger_change=True)

    # Optional fields
    fill_field(driver, f"{prefix}link", DATA_FILE_EXAMPLE["link"], trigger_change=True)
    fill_field(
        driver, f"{prefix}description", DATA_FILE_EXAMPLE["description"], trigger_change=True
    )
    fill_field(driver, f"{prefix}version", DATA_FILE_EXAMPLE["version"], trigger_change=True)
    fill_field(driver, f"{prefix}file_type", DATA_FILE_EXAMPLE["file_type"], trigger_change=True)


def expand_optional_fields(driver):
    """Expand the optional fields section if collapsed."""
    try:
        toggle = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "[data-testid='section-optional-toggle']")
            )
        )
        # Check if collapsible content is visible
        parent = toggle.find_element(By.XPATH, "..")
        content = parent.find_element(By.CSS_SELECTOR, ".collapsible-content")
        if not content.is_displayed():
            toggle.click()
            time.sleep(CLICK_DELAY)
            # Wait for content to be visible
            WebDriverWait(driver, 5).until(EC.visibility_of(content))
    except Exception:
        pass  # Optional fields section may not exist


@pytest.mark.ui
class TestCreateInvestigationRequiredFields:
    """Test creating an Investigation with required fields only."""

    def test_create_investigation_required_fields(self, browser):
        """Create Investigation with only required fields using YAML examples."""
        # Navigate to home
        browser.get(BASE_URL)
        time.sleep(1)  # Wait for page to fully load

        # Click "+ Investigation" button
        start_new_investigation(browser)

        # Verify form is displayed
        assert element_exists(browser, "form-entity")
        assert element_exists(browser, "input-unique-id")

        # Fill required fields from YAML example
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"])
        fill_field(browser, "input-title", INV_EXAMPLE["title"])

        # Click Create button
        click_button(browser, "btn-create")

        # Verify success message appears
        time.sleep(CLICK_DELAY)
        assert element_exists(browser, "form-success")
        body_text = browser.find_element(By.TAG_NAME, "body").text
        assert "Created Investigation" in body_text

        # Verify form switched to edit mode (Update button visible)
        assert element_exists(browser, "btn-update")


@pytest.mark.ui
class TestCreateInvestigationAllFields:
    """Test creating an Investigation with all fields using YAML examples."""

    def test_create_investigation_all_fields(self, browser):
        """Create Investigation with all fields from YAML examples including nested entities."""
        # Navigate to home
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Click "+ Investigation" button
        start_new_investigation(browser)

        # Fill required fields from YAML example
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"])
        fill_field(browser, "input-title", INV_EXAMPLE["title"])

        # Expand optional fields
        expand_optional_fields(browser)

        # Fill all optional scalar fields from YAML example
        # Note: miappe_version is auto-populated and readonly, so we skip it
        fill_field(browser, "input-description", INV_EXAMPLE["description"])
        fill_field(browser, "input-submission-date", INV_EXAMPLE["submission_date"])
        fill_field(browser, "input-public-release-date", INV_EXAMPLE["public_release_date"])
        fill_field(browser, "input-license", INV_EXAMPLE["license"])

        # Fill associated_publications (textarea, one per line)
        pubs = INV_EXAMPLE.get("associated_publications", [])
        if pubs:
            fill_field(browser, "input-associated-publications", "\n".join(pubs))

        # Click Create button
        click_button(browser, "btn-create")
        time.sleep(1)

        # After creation, we're automatically in edit mode
        assert element_exists(browser, "btn-update")

        # Add contact from YAML Person example - fill ALL fields
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-contacts")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_person_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Add study from YAML Study example - fill ALL fields
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_study_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Verify Investigation form shows correct counts
        expand_optional_fields(browser)
        contacts_btn = browser.find_element(By.CSS_SELECTOR, "[data-testid='btn-nested-contacts']")
        assert "(1)" in contacts_btn.text, f"Expected 1 contact, got: {contacts_btn.text}"

        studies_btn = browser.find_element(By.CSS_SELECTOR, "[data-testid='btn-nested-studies']")
        assert "(1)" in studies_btn.text, f"Expected 1 study, got: {studies_btn.text}"

        # Update the Investigation to save all nested data
        click_button(browser, "btn-update")
        time.sleep(CLICK_DELAY)

        # Verify we're still in edit mode after update
        assert element_exists(browser, "btn-update")


@pytest.mark.ui
class TestAutoPopulatedFields:
    """Test auto-populated fields like miappe_version."""

    def test_miappe_version_auto_populated(self, browser):
        """Verify miappe_version is auto-populated and readonly."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        start_new_investigation(browser)

        # Expand optional fields to see miappe_version
        expand_optional_fields(browser)

        # Find miappe_version input
        version_input = browser.find_element(
            By.CSS_SELECTOR, "[data-testid='input-miappe-version']"
        )

        # Verify it has a value (auto-populated)
        assert (
            version_input.get_attribute("value") == "1.1"
        ), f"Expected miappe_version to be '1.1', got: {version_input.get_attribute('value')}"

        # Verify it is readonly
        assert (
            version_input.get_attribute("readonly") is not None
        ), "miappe_version field should be readonly"

        # Verify it has the readonly class
        assert "form-input-readonly" in version_input.get_attribute(
            "class"
        ), "miappe_version field should have form-input-readonly class"

    def test_date_format_iso8601(self, browser):
        """Verify date fields use ISO 8601 format (YYYY-MM-DD)."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        start_new_investigation(browser)

        # Expand optional fields to see date fields
        expand_optional_fields(browser)

        # Find a date input
        date_input = browser.find_element(By.CSS_SELECTOR, "[data-testid='input-submission-date']")

        # Verify it uses text type with pattern
        assert (
            date_input.get_attribute("type") == "text"
        ), "Date field should be type='text' for consistent formatting"

        pattern = date_input.get_attribute("pattern")
        assert (
            pattern == r"\d{4}-\d{2}-\d{2}"
        ), f"Date field should have ISO 8601 pattern, got: {pattern}"

        placeholder = date_input.get_attribute("placeholder")
        assert (
            placeholder == "YYYY-MM-DD"
        ), f"Date field placeholder should be 'YYYY-MM-DD', got: {placeholder}"

    def test_table_cell_invalid_latitude_highlighted(self, browser):
        """Verify invalid latitude values get highlighted in table cells."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create Investigation first
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", "lat-test-inv")
        fill_field(browser, "input-title", "Latitude Test Investigation")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Add a study
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        # Fill required study fields
        fill_field(browser, "cell-0-unique_id", "lat-test-study", trigger_change=True)
        fill_field(browser, "cell-0-title", "Latitude Test Study", trigger_change=True)

        # Enter invalid latitude (43333 is way out of range, valid is -90 to 90)
        lat_input = browser.find_element(By.CSS_SELECTOR, "[data-testid='cell-0-latitude']")
        lat_input.clear()
        lat_input.send_keys("43333")
        # Trigger input event to invoke validation
        browser.execute_script(
            "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", lat_input
        )
        time.sleep(0.5)

        # Verify the input has the 'invalid' class (red background)
        classes = lat_input.get_attribute("class")
        assert "invalid" in classes, f"Invalid latitude should be highlighted. Classes: {classes}"

        # Now enter a valid latitude
        lat_input.clear()
        lat_input.send_keys("52.5")
        browser.execute_script(
            "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", lat_input
        )
        time.sleep(0.5)

        # Verify the 'invalid' class is removed
        classes = lat_input.get_attribute("class")
        assert (
            "invalid" not in classes
        ), f"Valid latitude should not be highlighted. Classes: {classes}"


@pytest.mark.ui
class TestAddNestedContacts:
    """Test adding nested contacts to an Investigation using YAML examples."""

    def test_add_nested_contacts(self, browser):
        """Create Investigation and add contacts from YAML Person example."""
        # Navigate to home
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Click "+ Investigation" button
        start_new_investigation(browser)

        # Fill required fields from YAML example
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"] + "-contacts")
        fill_field(browser, "input-title", INV_EXAMPLE["title"] + " (Contacts Test)")

        # Click Create button
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Expand optional fields to access contacts button
        expand_optional_fields(browser)

        # Click on contacts nested field button
        click_button(browser, "btn-nested-contacts")

        # Verify table view is displayed
        assert element_exists(browser, "table-add-row")
        assert element_exists(browser, "table-back")

        # Click "+ Add Row" button
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        # Fill ALL contact fields from YAML Person example
        fill_all_person_fields(browser)

        # Click "Save & Back" button
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Verify we're back on the Investigation form
        assert element_exists(browser, "form-entity")

        # Expand optional fields to check contacts count
        expand_optional_fields(browser)

        # Verify contacts button shows count of 1
        contacts_btn = browser.find_element(By.CSS_SELECTOR, "[data-testid='btn-nested-contacts']")
        assert "(1)" in contacts_btn.text


@pytest.mark.ui
class TestEditInvestigation:
    """Test editing an existing Investigation."""

    def test_edit_investigation(self, browser):
        """Edit an existing Investigation and verify changes persist."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create an Investigation first
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", "INV-EDIT-001")
        fill_field(browser, "input-title", "Original Title")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Modify the title
        title_field = browser.find_element(By.CSS_SELECTOR, "[data-testid='input-title']")
        title_field.clear()
        title_field.send_keys("Updated Title")
        time.sleep(FILL_DELAY)

        # Click Update
        click_button(browser, "btn-update")
        time.sleep(CLICK_DELAY)

        # Verify success message appears
        body_text = browser.find_element(By.TAG_NAME, "body").text
        assert "Updated" in body_text or element_exists(browser, "form-success")


@pytest.mark.ui
@pytest.mark.skip(reason="Delete functionality not available in current UI layout")
class TestDeleteInvestigation:
    """Test deleting an Investigation."""

    def test_delete_investigation(self, browser):
        """Delete an Investigation and verify it's removed."""
        pass  # Skipped - no delete button in current layout


@pytest.mark.ui
class TestAddNestedStudy:
    """Test adding a nested Study to an Investigation using YAML examples."""

    def test_add_nested_study(self, browser):
        """Add a Study from YAML example to an Investigation through nested table."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create Investigation using YAML example
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"] + "-study")
        fill_field(browser, "input-title", INV_EXAMPLE["title"] + " (Study Test)")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Expand optional fields
        expand_optional_fields(browser)

        # Click on studies nested field button
        click_button(browser, "btn-nested-studies")

        # Add a row
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        # Fill ALL study fields from YAML Study example
        fill_all_study_fields(browser)

        # Save & Back
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Verify studies count shows 1
        expand_optional_fields(browser)
        studies_btn = browser.find_element(By.CSS_SELECTOR, "[data-testid='btn-nested-studies']")
        assert "(1)" in studies_btn.text


@pytest.mark.ui
class TestValidation:
    """Test validation functionality using the Validate button."""

    def test_validate_valid_investigation(self, browser):
        """Validate a correctly filled Investigation form.

        Note: Full validation requires contacts and studies (cardinality rules).
        This test validates field-level requirements (unique_id, title) pass.
        """
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create new Investigation form
        start_new_investigation(browser)

        # Fill required fields from YAML example
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"])
        fill_field(browser, "input-title", INV_EXAMPLE["title"])

        # Create the investigation first
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")
        time.sleep(CLICK_DELAY)

        # Add a contact (required by cardinality rule)
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-contacts")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)
        fill_field(browser, "cell-0-name", "Test Contact", trigger_change=True)
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Add a study (required by cardinality rule)
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)
        fill_field(browser, "cell-0-unique_id", "STU-VAL-001", trigger_change=True)
        fill_field(browser, "cell-0-title", "Validation Test Study", trigger_change=True)
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Click Validate button
        click_button(browser, "btn-validate")

        # Verify validation success message
        assert element_exists(browser, "validation-success")
        success_msg = browser.find_element(By.CSS_SELECTOR, "[data-testid='validation-success']")
        assert "Validation passed" in success_msg.text

    def test_validate_missing_required_fields(self, browser):
        """Validate form with missing required fields shows errors."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create new Investigation form
        start_new_investigation(browser)

        # Fill only unique_id, leave title empty (both required)
        fill_field(browser, "input-unique-id", "INV-VAL-001")

        # Click Validate button
        click_button(browser, "btn-validate")

        # Verify validation errors are displayed
        assert element_exists(browser, "validation-errors")
        errors_div = browser.find_element(By.CSS_SELECTOR, "[data-testid='validation-errors']")
        assert "Validation failed" in errors_div.text

        # Verify specific field error for title
        assert element_exists(browser, "validation-error-title")
        title_error = browser.find_element(
            By.CSS_SELECTOR, "[data-testid='validation-error-title']"
        )
        assert "title" in title_error.text.lower()

    def test_validate_invalid_pattern(self, browser):
        """Validate form with invalid field pattern shows error."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create new Investigation form
        start_new_investigation(browser)

        # Fill required fields
        fill_field(browser, "input-unique-id", "INVALID ID WITH SPACES!")  # Invalid pattern
        fill_field(browser, "input-title", "Test Investigation")

        # Click Validate button
        click_button(browser, "btn-validate")

        # Verify validation errors for pattern violation
        assert element_exists(browser, "validation-errors")
        assert element_exists(browser, "validation-error-unique_id")

    def test_validate_then_fix_and_revalidate(self, browser):
        """Validate with errors, fix them, and revalidate successfully.

        Tests the workflow of fixing validation errors iteratively.
        """
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create new Investigation form
        start_new_investigation(browser)

        # Fill only unique_id (missing title)
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"])

        # Validate - should fail (missing title)
        click_button(browser, "btn-validate")
        assert element_exists(browser, "validation-errors")

        # Fix by adding title
        fill_field(browser, "input-title", INV_EXAMPLE["title"])

        # Create the investigation
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Validate - should still fail (missing contacts and studies)
        click_button(browser, "btn-validate")
        assert element_exists(browser, "validation-errors")

        # Add a contact
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-contacts")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)
        fill_field(browser, "cell-0-name", "Fix Test Contact", trigger_change=True)
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Add a study
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)
        fill_field(browser, "cell-0-unique_id", "STU-FIX-001", trigger_change=True)
        fill_field(browser, "cell-0-title", "Fix Test Study", trigger_change=True)
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Validate again - should pass now
        click_button(browser, "btn-validate")
        assert element_exists(browser, "validation-success")

    def test_html5_validation_required_fields(self, browser):
        """Verify HTML5 required attribute is set on required fields."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create new Investigation form
        start_new_investigation(browser)

        # Check if unique_id field has required attribute
        unique_id = browser.find_element(By.CSS_SELECTOR, "[data-testid='input-unique-id']")
        assert unique_id.get_attribute("required") is not None

        # Check if title field has required attribute
        title = browser.find_element(By.CSS_SELECTOR, "[data-testid='input-title']")
        assert title.get_attribute("required") is not None


@pytest.mark.ui
@pytest.mark.skip(
    reason="Profile switching via dropdown removed from UI - profile selected at Investigation creation"
)
class TestProfileSwitch:
    """Test profile switching between miappe and isa."""

    def test_switch_profile_clears_state(self, browser):
        """Switch profile and verify state is cleared."""
        pass  # Skipped - profile is now selected during Investigation creation


@pytest.mark.ui
class TestNestedEntityEditing:
    """Test editing nested entities by clicking on table rows using YAML examples."""

    def test_edit_nested_study(self, browser):
        """Click a Study row in table to open edit form for that Study."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create Investigation with a Study using YAML examples
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"] + "-nested")
        fill_field(browser, "input-title", INV_EXAMPLE["title"] + " (Nested Edit)")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")
        time.sleep(CLICK_DELAY)

        # Expand optional fields
        expand_optional_fields(browser)

        # Click on studies nested field button
        click_button(browser, "btn-nested-studies")

        # Add a Study row from YAML example - fill ALL fields
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_study_fields(browser)

        # Save the Study first to ensure it's persisted
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Go back to Studies table
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        time.sleep(CLICK_DELAY)

        # Click on the Study row's Edit button to edit it
        click_element(browser, "row-edit-0")
        time.sleep(CLICK_DELAY)

        # Verify we're now in Study edit form
        # Should show Study-specific fields
        assert element_exists(browser, "form-entity")

        # Verify breadcrumb shows navigation path
        assert element_exists(browser, "breadcrumb")
        breadcrumb = browser.find_element(By.CSS_SELECTOR, "[data-testid='breadcrumb']")
        # Breadcrumb shows: Investigation title > Study title
        assert "(Nested Edit)" in breadcrumb.text  # Part of Investigation title
        assert STUDY_EXAMPLE["title"] in breadcrumb.text  # Study title

        # Verify the Study form has the correct values from YAML example
        unique_id_field = browser.find_element(By.CSS_SELECTOR, "[data-testid='input-unique-id']")
        assert unique_id_field.get_attribute("value") == STUDY_EXAMPLE["unique_id"]

        # Verify back button exists
        assert element_exists(browser, "btn-cancel")

        # Click back button to return to table
        click_button(browser, "btn-cancel")
        time.sleep(CLICK_DELAY)

        # Verify we're back on the table view
        assert element_exists(browser, "table-back")

    def test_deep_nesting_navigation(self, browser):
        """Navigate from Investigation > Study > BiologicalMaterial using YAML examples."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create Investigation using YAML examples
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"] + "-deep")
        fill_field(browser, "input-title", INV_EXAMPLE["title"] + " (Deep Nesting)")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Expand optional fields and add a Study from YAML example - fill ALL fields
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_study_fields(browser)

        # Save and go back to Investigation
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Go back to Studies table
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        time.sleep(CLICK_DELAY)

        # Click on Study row to edit it
        click_element(browser, "row-edit-0")
        time.sleep(CLICK_DELAY)

        # Verify we're in Study form
        assert element_exists(browser, "form-entity")

        # Expand optional fields in Study form
        expand_optional_fields(browser)

        # Click on biological_materials nested field
        click_button(browser, "btn-nested-biological-materials")
        time.sleep(CLICK_DELAY)

        # Verify we're in BiologicalMaterial table view
        assert element_exists(browser, "table-add-row")

        # Add a BiologicalMaterial from YAML example - fill ALL fields
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_biological_material_fields(browser)

        # Save
        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Verify we're back in Study form
        assert element_exists(browser, "form-entity")

        # Go back to biological_materials table
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-biological-materials")
        time.sleep(CLICK_DELAY)

        # Click on BiologicalMaterial row to edit
        click_element(browser, "row-edit-0")
        time.sleep(CLICK_DELAY)

        # Verify breadcrumb shows full path
        assert element_exists(browser, "breadcrumb")
        breadcrumb = browser.find_element(By.CSS_SELECTOR, "[data-testid='breadcrumb']")
        breadcrumb_text = breadcrumb.text
        # Breadcrumb shows: Investigation title > Study title > BiologicalMaterial unique_id
        assert "(Deep Nesting)" in breadcrumb_text  # Part of Investigation title
        assert STUDY_EXAMPLE["title"] in breadcrumb_text  # Study title
        assert BIO_MAT_EXAMPLE["unique_id"] in breadcrumb_text  # BiologicalMaterial id

        # Verify BiologicalMaterial form fields from YAML example
        unique_id_field = browser.find_element(By.CSS_SELECTOR, "[data-testid='input-unique-id']")
        assert unique_id_field.get_attribute("value") == BIO_MAT_EXAMPLE["unique_id"]

        # Deep nesting was successful - we can navigate to a BiologicalMaterial
        # nested within a Study nested within an Investigation


@pytest.mark.ui
class TestCreateAllEntityTypes:
    """Test creating ALL entity types from YAML examples in one comprehensive test."""

    def test_create_all_entities(self, browser):
        """Create Investigation with ALL nested entity types from YAML examples.

        This test creates one instance of 10 MIAPPE entity types:
        - Investigation (root)
        - Person (via contacts)
        - Study (via studies)
        - DataFile (via Study > data_files)
        - BiologicalMaterial (via Study > biological_materials)
        - ObservationUnit (via Study > observation_units)
        - ObservedVariable (via Study > observed_variables)
        - Factor (via Study > factors)
        - Event (via Study > events)
        - Environment (via Study > environments)
        """
        browser.get(BASE_URL)
        time.sleep(1)

        # === 1. Create Investigation with all fields ===
        start_new_investigation(browser)

        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"])
        fill_field(browser, "input-title", INV_EXAMPLE["title"])

        expand_optional_fields(browser)
        fill_field(browser, "input-description", INV_EXAMPLE["description"])
        fill_field(browser, "input-submission-date", INV_EXAMPLE["submission_date"])
        fill_field(browser, "input-public-release-date", INV_EXAMPLE["public_release_date"])
        fill_field(browser, "input-license", INV_EXAMPLE["license"])
        # Note: miappe_version is auto-populated and readonly, so we skip it

        pubs = INV_EXAMPLE.get("associated_publications", [])
        if pubs:
            fill_field(browser, "input-associated-publications", "\n".join(pubs))

        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # Click on created Investigation
        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # === 2. Add Person (via contacts) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-contacts")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_person_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # === 3. Add Study ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_study_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Navigate into Study to add nested entities
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        time.sleep(CLICK_DELAY)
        click_element(browser, "row-edit-0")
        time.sleep(CLICK_DELAY)

        # === 4. Add DataFile (via Study > data_files) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-data-files")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_data_file_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # === 5. Add BiologicalMaterial (via Study > biological_materials) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-biological-materials")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_biological_material_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # === 6. Add ObservationUnit (via Study > observation_units) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-observation-units")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_observation_unit_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # === 7. Add ObservedVariable (via Study > observed_variables) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-observed-variables")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_observed_variable_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # === 8. Add Factor (via Study > factors) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-factors")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_factor_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # === 9. Add Event (via Study > events) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-events")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_event_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # === 10. Add Environment (via Study > environments) ===
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-environments")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_all_environment_fields(browser)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Verify we're on Study form after adding all nested entities
        assert element_exists(browser, "form-entity")

        # Navigate back to Investigation
        click_button(browser, "btn-cancel")  # Study form to studies table
        time.sleep(CLICK_DELAY)
        click_button(browser, "table-back")  # studies table to Investigation form
        time.sleep(CLICK_DELAY)

        # Verify Investigation form is displayed
        assert element_exists(browser, "form-entity")
        assert element_exists(browser, "btn-update")

        # All 10 entity types were successfully created:
        # Investigation, Person, Study, DataFile, BiologicalMaterial, ObservationUnit,
        # ObservedVariable, Factor, Event, Environment


@pytest.mark.ui
class TestExcelExport:
    """Test Excel export functionality."""

    def test_export_button_visible_in_edit_mode(self, browser):
        """Export button should be visible when editing an entity."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create an Investigation
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", "INV-EXPORT-001")
        fill_field(browser, "input-title", "Export Test Investigation")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Verify export button is visible in edit mode
        assert element_exists(browser, "btn-export")

    def test_export_button_not_in_create_mode(self, browser):
        """Export button should not be visible when creating a new entity."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Open new Investigation form
        start_new_investigation(browser)

        # Verify export button is NOT visible in create mode
        assert not element_exists(browser, "btn-export")

    def test_export_downloads_file(self, browser):
        """Click export button and verify it triggers download."""
        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create an Investigation
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", INV_EXAMPLE["unique_id"])
        fill_field(browser, "input-title", INV_EXAMPLE["title"])
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Get export button href
        export_btn = browser.find_element(By.CSS_SELECTOR, "[data-testid='btn-export']")
        export_url = export_btn.get_attribute("href")

        # Verify export URL is correct
        assert "/export" in export_url

        # Test export endpoint directly using requests
        import urllib.request

        req = urllib.request.Request(f"{BASE_URL}/export")
        response = urllib.request.urlopen(req, timeout=10)

        # Verify response headers
        content_type = response.headers.get("Content-Type")
        assert "spreadsheetml" in content_type or "excel" in content_type.lower()

        content_disp = response.headers.get("Content-Disposition")
        assert "attachment" in content_disp
        assert ".xlsx" in content_disp

        # Verify content is valid Excel (starts with PK for ZIP format)
        content = response.read()
        assert content[:2] == b"PK"  # Excel files are ZIP archives

    def test_export_contains_investigation_data(self, browser):
        """Verify exported Excel contains Investigation data."""
        from openpyxl import load_workbook

        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create Investigation with specific data
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", "INV-EXPORT-DATA-001")
        fill_field(browser, "input-title", "Export Data Test")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # Download export
        import urllib.request
        from io import BytesIO

        req = urllib.request.Request(f"{BASE_URL}/export")
        response = urllib.request.urlopen(req, timeout=10)
        content = response.read()

        # Load workbook and verify content
        wb = load_workbook(BytesIO(content))

        # Verify Investigation sheet exists
        assert "Investigation" in wb.sheetnames

        # Verify data is in the sheet
        ws = wb["Investigation"]
        rows = list(ws.values)
        headers = rows[0]
        data_rows = rows[1:]

        # Verify headers include key fields
        assert "unique_id" in headers
        assert "title" in headers

        # Verify at least one data row with our values
        found = False
        for row in data_rows:
            row_dict = dict(zip(headers, row, strict=False))
            if row_dict.get("unique_id") == "INV-EXPORT-DATA-001":
                found = True
                assert row_dict.get("title") == "Export Data Test"
                break

        assert found, "Investigation data not found in export"

    def test_export_contains_nested_entities(self, browser):
        """Verify exported Excel contains nested entities (Study, Person)."""
        from openpyxl import load_workbook

        browser.get(BASE_URL)
        time.sleep(CLICK_DELAY)

        # Create Investigation
        start_new_investigation(browser)
        fill_field(browser, "input-unique-id", "INV-NESTED-EXPORT-001")
        fill_field(browser, "input-title", "Nested Export Test")
        click_button(browser, "btn-create")
        time.sleep(CLICK_DELAY)

        # After creation, we're already in edit mode
        assert element_exists(browser, "btn-update")

        # Add a contact (Person)
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-contacts")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_field(browser, "cell-0-name", "Export Test Person", trigger_change=True)
        fill_field(browser, "cell-0-email", "export@test.com", trigger_change=True)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Add a study
        expand_optional_fields(browser)
        click_button(browser, "btn-nested-studies")
        click_button(browser, "table-add-row")
        time.sleep(CLICK_DELAY)

        fill_field(browser, "cell-0-unique_id", "STU-EXPORT-001", trigger_change=True)
        fill_field(browser, "cell-0-title", "Export Test Study", trigger_change=True)

        click_button(browser, "table-back")
        time.sleep(CLICK_DELAY)

        # Save the investigation
        click_button(browser, "btn-update")
        time.sleep(CLICK_DELAY)

        # Download export
        import urllib.request
        from io import BytesIO

        req = urllib.request.Request(f"{BASE_URL}/export")
        response = urllib.request.urlopen(req, timeout=10)
        content = response.read()

        # Load workbook and verify all entity sheets exist
        wb = load_workbook(BytesIO(content))

        # Verify Investigation sheet
        assert (
            "Investigation" in wb.sheetnames
        ), f"Missing Investigation sheet. Sheets: {wb.sheetnames}"

        # Verify Person sheet (contacts)
        assert "Person" in wb.sheetnames, f"Missing Person sheet. Sheets: {wb.sheetnames}"

        # Verify Study sheet
        assert "Study" in wb.sheetnames, f"Missing Study sheet. Sheets: {wb.sheetnames}"

        # Verify Person data
        ws_person = wb["Person"]
        person_rows = list(ws_person.values)
        person_headers = person_rows[0]
        person_data = person_rows[1:] if len(person_rows) > 1 else []

        assert "name" in person_headers
        found_person = False
        for row in person_data:
            row_dict = dict(zip(person_headers, row, strict=False))
            if row_dict.get("name") == "Export Test Person":
                found_person = True
                assert row_dict.get("email") == "export@test.com"
                break
        assert found_person, "Person data not found in export"

        # Verify Study data
        ws_study = wb["Study"]
        study_rows = list(ws_study.values)
        study_headers = study_rows[0]
        study_data = study_rows[1:] if len(study_rows) > 1 else []

        assert "unique_id" in study_headers
        found_study = False
        for row in study_data:
            row_dict = dict(zip(study_headers, row, strict=False))
            if row_dict.get("unique_id") == "STU-EXPORT-001":
                found_study = True
                assert row_dict.get("title") == "Export Test Study"
                break
        assert found_study, "Study data not found in export"
