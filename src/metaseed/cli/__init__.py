"""CLI module for Metaseed."""

from pathlib import Path
from typing import Annotated

import typer
import yaml

from metaseed import __version__
from metaseed.cli.output import CheckOutput, echo_error, echo_success
from metaseed.importers import ISAImporter
from metaseed.models import get_model
from metaseed.profiles import ProfileFactory
from metaseed.specs.loader import SpecLoader, SpecLoadError
from metaseed.storage import JsonStorage, StorageError, YamlStorage
from metaseed.validators import DatasetValidator
from metaseed.validators import validate as validate_data

app = typer.Typer(
    name="metaseed",
    help="Tools for creating, editing, and validating experimental metadata following MIAPPE standards.",
    no_args_is_help=True,
)

# Exit codes
EXIT_SUCCESS = 0
EXIT_VALIDATION_ERROR = 1
EXIT_INPUT_ERROR = 2
EXIT_CONFIG_ERROR = 3


def resolve_profile_version(profile: str | None, version: str | None) -> tuple[str, str]:
    """Resolve profile and version with smart defaults.

    Args:
        profile: Profile name, or None for default.
        version: Version string, or None for latest.

    Returns:
        Tuple of (profile, version) with defaults resolved.

    Raises:
        typer.Exit: If profile is unknown (exit code 3).
    """
    factory = ProfileFactory()

    if profile is None:
        profile = factory.get_default_profile()

    if profile not in factory.list_profiles():
        echo_error(f"Unknown profile '{profile}'")
        raise typer.Exit(EXIT_CONFIG_ERROR)

    if version is None:
        latest = factory.get_latest_version(profile)
        if latest is None:
            echo_error(f"No versions found for profile '{profile}'")
            raise typer.Exit(EXIT_CONFIG_ERROR)
        version = latest

    return profile, version


@app.command()
def version() -> None:
    """Show the version."""
    typer.echo(f"metaseed {__version__}")


@app.command()
def profiles(
    verbose: Annotated[
        bool, typer.Option("--verbose", "-v", help="Show detailed information")
    ] = False,
) -> None:
    """List available profiles and versions."""
    factory = ProfileFactory()
    profile_list = factory.get_profile_info()

    if not profile_list:
        typer.echo("No profiles available.")
        return

    if verbose:
        for info in profile_list:
            typer.echo(f"{info['name']}:")
            typer.echo(f"  versions: {', '.join(info['versions'])}")
            typer.echo(f"  latest: {info['latest']}")
    else:
        default = factory.get_default_profile()
        for info in profile_list:
            marker = " (default)" if info["name"] == default else ""
            typer.echo(f"  {info['name']}{marker}")


@app.command()
def validate(
    file: Annotated[Path, typer.Argument(help="Path to the file to validate")],
    entity: Annotated[str, typer.Option("--entity", "-e", help="Entity type")] = "investigation",
    profile: Annotated[str | None, typer.Option("--profile", "-p", help="Profile name")] = None,
    version: Annotated[str | None, typer.Option("--version", "-v", help="Profile version")] = None,
) -> None:
    """Validate a metadata file against a profile."""
    profile, version = resolve_profile_version(profile, version)

    if not file.exists():
        echo_error(f"File not found: {file}")
        raise typer.Exit(EXIT_INPUT_ERROR)

    try:
        content = file.read_text(encoding="utf-8")
        data = yaml.safe_load(content)
        if data is None:
            data = {}
    except yaml.YAMLError as e:
        echo_error(f"Invalid YAML: {e}")
        raise typer.Exit(EXIT_INPUT_ERROR) from None

    errors = validate_data(data, entity, version, profile=profile)

    if errors:
        typer.echo(f"Validation failed with {len(errors)} error(s):")
        for error in errors:
            typer.echo(f"  - {error.field}: {error.message}")
        raise typer.Exit(EXIT_VALIDATION_ERROR)
    echo_success(f"Validation passed. File is valid {entity} ({profile} v{version}).")


@app.command()
def template(
    entity: Annotated[str, typer.Argument(help="Entity type to generate template for")],
    output: Annotated[Path | None, typer.Option("--output", "-o", help="Output file path")] = None,
    format: Annotated[str, typer.Option("--format", "-f", help="Output format")] = "yaml",
    profile: Annotated[str | None, typer.Option("--profile", "-p", help="Profile name")] = None,
    version: Annotated[str | None, typer.Option("--version", "-v", help="Profile version")] = None,
) -> None:
    """Generate a template file for an entity."""
    profile, version = resolve_profile_version(profile, version)

    try:
        loader = SpecLoader(profile=profile)
        spec = loader.load_entity(entity.lower(), version)
    except SpecLoadError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_CONFIG_ERROR) from None

    # Build template with empty/example values
    template_data = {}
    for field in spec.fields:
        if field.required:
            if field.type.value == "string":
                template_data[field.name] = f"<{field.name}>"
            elif field.type.value == "integer":
                template_data[field.name] = 0
            elif field.type.value == "float":
                template_data[field.name] = 0.0
            elif field.type.value == "boolean":
                template_data[field.name] = False
            elif field.type.value == "date":
                template_data[field.name] = "2024-01-01"
            elif field.type.value == "datetime":
                template_data[field.name] = "2024-01-01T00:00:00"
            elif field.type.value == "list":
                template_data[field.name] = []
            else:
                template_data[field.name] = None
        else:
            # Add commented example for optional fields
            template_data[f"# {field.name}"] = None

    # Generate output
    if format.lower() == "json":
        import json

        content = json.dumps(template_data, indent=2)
    else:
        content = yaml.dump(template_data, default_flow_style=False, sort_keys=False)

    if output:
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(content, encoding="utf-8")
        typer.echo(f"Template written to {output}")
    else:
        typer.echo(content)


@app.command()
def convert(
    input_file: Annotated[Path, typer.Argument(help="Input file path")],
    output_file: Annotated[Path, typer.Argument(help="Output file path")],
    entity: Annotated[str, typer.Option("--entity", "-e", help="Entity type")] = "investigation",
    profile: Annotated[str | None, typer.Option("--profile", "-p", help="Profile name")] = None,
    version: Annotated[str | None, typer.Option("--version", "-v", help="Profile version")] = None,
) -> None:
    """Convert between YAML and JSON formats."""
    profile, version = resolve_profile_version(profile, version)

    if not input_file.exists():
        echo_error(f"File not found: {input_file}")
        raise typer.Exit(EXIT_INPUT_ERROR)

    try:
        Model = get_model(entity, version)
    except SpecLoadError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_CONFIG_ERROR) from None

    # Determine input format
    input_suffix = input_file.suffix.lower()
    if input_suffix in [".yaml", ".yml"]:
        input_storage = YamlStorage()
    elif input_suffix == ".json":
        input_storage = JsonStorage()
    else:
        typer.echo(f"Error: Unknown input format: {input_suffix}", err=True)
        raise typer.Exit(1)

    # Determine output format
    output_suffix = output_file.suffix.lower()
    if output_suffix in [".yaml", ".yml"]:
        output_storage = YamlStorage()
    elif output_suffix == ".json":
        output_storage = JsonStorage()
    else:
        typer.echo(f"Error: Unknown output format: {output_suffix}", err=True)
        raise typer.Exit(1)

    try:
        entity_instance = input_storage.load(input_file, Model)
        output_storage.save(entity_instance, output_file)
        echo_success(f"Converted {input_file} to {output_file}")
    except StorageError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_INPUT_ERROR) from None


@app.command()
def entities(
    profile: Annotated[str | None, typer.Option("--profile", "-p", help="Profile name")] = None,
    version: Annotated[str | None, typer.Option("--version", "-v", help="Profile version")] = None,
) -> None:
    """List available entities for a profile."""
    profile, version = resolve_profile_version(profile, version)

    try:
        loader = SpecLoader(profile=profile)
        entity_list = loader.list_entities(version)
    except SpecLoadError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_CONFIG_ERROR) from None

    typer.echo(f"Available entities ({profile} v{version}):")
    for entity in sorted(entity_list):
        typer.echo(f"  - {entity}")


@app.command()
def check(
    path: Annotated[Path, typer.Argument(help="Path to file or directory to check")],
    profile: Annotated[str | None, typer.Option("--profile", "-p", help="Profile name")] = None,
    version: Annotated[str | None, typer.Option("--version", "-v", help="Profile version")] = None,
    verbose: Annotated[bool, typer.Option("--verbose", help="Show detailed information")] = False,
    quiet: Annotated[bool, typer.Option("--quiet", "-q", help="Suppress non-error output")] = False,
) -> None:
    """Validate dataset with reference integrity checking.

    Checks a file or directory for:
    - Entity structure validation
    - Required field presence
    - Reference integrity (cross-entity references exist)
    """
    profile, version = resolve_profile_version(profile, version)

    if not path.exists():
        echo_error(f"Path not found: {path}")
        raise typer.Exit(EXIT_INPUT_ERROR)

    validator = DatasetValidator(profile=profile, version=version)
    output_formatter = CheckOutput(verbose=verbose, quiet=quiet)

    result = validator.validate_file(path) if path.is_file() else validator.validate_directory(path)

    output_formatter.print_result(result)

    if not result.is_valid:
        raise typer.Exit(EXIT_VALIDATION_ERROR)


@app.command(name="import")
def import_isa(
    path: Annotated[Path, typer.Argument(help="Path to ISA-JSON file or ISA-Tab directory")],
    output: Annotated[Path | None, typer.Option("--output", "-o", help="Output directory")] = None,
    format: Annotated[str, typer.Option("--format", "-f", help="Output format")] = "yaml",
) -> None:
    """Import ISA-Tab or ISA-JSON data to MIAPPE format.

    Supports:
    - ISA-JSON: Single .json file
    - ISA-Tab: Directory containing i_*.txt, s_*.txt, a_*.txt files
    """
    importer = ISAImporter()

    try:
        if path.is_file() and path.suffix.lower() == ".json":
            result = importer.import_json(path)
            typer.echo(f"Imported ISA-JSON: {path.name}")
        elif path.is_dir():
            result = importer.import_tab(path)
            typer.echo(f"Imported ISA-Tab: {path.name}")
        else:
            typer.echo("Error: Path must be a .json file or directory", err=True)
            raise typer.Exit(1)
    except Exception as e:
        typer.echo(f"Error importing: {e}", err=True)
        raise typer.Exit(1) from None

    # Print summary
    typer.echo(result.summary)

    # Show warnings if any
    for warning in result.warnings:
        typer.echo(f"  Warning: {warning}")

    # Output results
    if output:
        output.mkdir(parents=True, exist_ok=True)

        if format.lower() == "json":
            import json

            (output / "investigation.json").write_text(json.dumps(result.investigation, indent=2))
            for i, study in enumerate(result.studies):
                (output / f"study_{i+1}.json").write_text(json.dumps(study, indent=2))
        else:
            (output / "investigation.yaml").write_text(
                yaml.dump(result.investigation, default_flow_style=False)
            )
            for i, study in enumerate(result.studies):
                (output / f"study_{i+1}.yaml").write_text(
                    yaml.dump(study, default_flow_style=False)
                )

        typer.echo(f"Output written to: {output}")
    else:
        # Print investigation to stdout
        typer.echo("\n--- Investigation ---")
        typer.echo(yaml.dump(result.investigation, default_flow_style=False))

        if result.studies:
            typer.echo("--- Studies ---")
            for study in result.studies:
                typer.echo(yaml.dump(study, default_flow_style=False))


@app.command(name="ui")
def web_ui(
    host: Annotated[str, typer.Option("--host", "-h", help="Host to bind to")] = "127.0.0.1",
    port: Annotated[int, typer.Option("--port", "-p", help="Port to bind to")] = 8080,
) -> None:
    """Launch the web interface."""
    from metaseed.ui import run_ui

    typer.echo(f"Starting Metaseed web interface at http://{host}:{port}")
    run_ui(host=host, port=port)


@app.command()
def example(
    profile: Annotated[
        str | None, typer.Argument(help="Profile name (miappe, isa, isa-miappe-combined)")
    ] = None,
    output: Annotated[
        Path | None, typer.Option("--output", "-o", help="Output file path (.xlsx or .yaml)")
    ] = None,
    list_examples: Annotated[
        bool, typer.Option("--list", "-l", help="List available examples")
    ] = False,
) -> None:
    """Export example data for a profile.

    Examples are fully populated sample datasets demonstrating each profile's
    entity types and relationships.

    Usage:
        metaseed example --list              # List available examples
        metaseed example miappe -o out.xlsx  # Export MIAPPE example to Excel
        metaseed example isa -o out.yaml     # Export ISA example to YAML
    """
    import importlib.resources

    # Find examples directory
    try:
        examples_dir = Path(importlib.resources.files("metaseed")).parent.parent / "examples"
        if not examples_dir.exists():
            # Try relative to package
            examples_dir = Path(__file__).parent.parent.parent.parent / "examples"
    except Exception:
        examples_dir = Path(__file__).parent.parent.parent.parent / "examples"

    if not examples_dir.exists():
        echo_error(f"Examples directory not found at {examples_dir}")
        raise typer.Exit(EXIT_CONFIG_ERROR)

    # Get available examples (profile/version/example.yaml structure)
    example_files = {}
    if examples_dir.exists():
        for profile_dir in examples_dir.iterdir():
            if profile_dir.is_dir() and not profile_dir.name.startswith("."):
                for version_dir in profile_dir.iterdir():
                    if version_dir.is_dir():
                        yaml_files = list(version_dir.glob("*.yaml"))
                        if yaml_files:
                            key = f"{profile_dir.name}/{version_dir.name}"
                            example_files[key] = yaml_files[0]

    if list_examples or profile is None:
        typer.echo("Available example datasets:")
        typer.echo("")
        for name, path in sorted(example_files.items()):
            typer.echo(f"  {name:30} {path.name}")
        typer.echo("")
        typer.echo("Usage: metaseed example <profile/version> -o output.xlsx")
        typer.echo("       metaseed example miappe/1.1 -o example.xlsx")
        return

    # Handle both "profile/version" and just "profile" (uses latest)
    profile_input = profile.lower()
    if "/" in profile_input:
        example_key = profile_input
    else:
        # Find latest version for profile
        matching = [k for k in example_files if k.startswith(f"{profile_input}/")]
        if not matching:
            echo_error(
                f"No examples for profile '{profile_input}'. Available: {', '.join(sorted(example_files.keys()))}"
            )
            raise typer.Exit(EXIT_CONFIG_ERROR)
        example_key = sorted(matching)[-1]  # Latest version

    if example_key not in example_files:
        echo_error(
            f"Example not found: '{example_key}'. Available: {', '.join(sorted(example_files.keys()))}"
        )
        raise typer.Exit(EXIT_CONFIG_ERROR)

    example_file = example_files[example_key]
    if not example_file.exists():
        echo_error(f"Example file not found: {example_file}")
        raise typer.Exit(EXIT_INPUT_ERROR)

    # Load example data
    try:
        data = yaml.safe_load(example_file.read_text(encoding="utf-8"))
    except yaml.YAMLError as e:
        echo_error(f"Invalid YAML in example: {e}")
        raise typer.Exit(EXIT_INPUT_ERROR) from None

    if output is None:
        # Print to stdout as YAML
        typer.echo(yaml.dump(data, default_flow_style=False, sort_keys=False, allow_unicode=True))
        return

    output_suffix = output.suffix.lower()

    if output_suffix in [".yaml", ".yml"]:
        # Copy YAML file
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(
            yaml.dump(data, default_flow_style=False, sort_keys=False, allow_unicode=True),
            encoding="utf-8",
        )
        echo_success(f"Example exported to {output}")

    elif output_suffix == ".xlsx":
        # Export to Excel
        try:
            _export_example_to_excel(data, output)
            echo_success(f"Example exported to {output}")
        except ImportError:
            echo_error("openpyxl is required for Excel export. Install with: pip install openpyxl")
            raise typer.Exit(EXIT_CONFIG_ERROR) from None

    elif output_suffix == ".json":
        import json

        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        echo_success(f"Example exported to {output}")

    else:
        echo_error(f"Unknown output format: {output_suffix}. Use .xlsx, .yaml, or .json")
        raise typer.Exit(EXIT_INPUT_ERROR)


def _export_example_to_excel(data: dict, output: Path) -> None:
    """Export example data to Excel with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    def flatten_entity(entity_data: dict, parent_fields: dict | None = None) -> dict:
        """Flatten nested entity, excluding nested lists."""
        flat = {}
        if parent_fields:
            flat.update(parent_fields)
        for key, value in entity_data.items():
            if not isinstance(value, list | dict) or key in [
                "associated_publications",
                "external_references",
            ]:
                flat[key] = (
                    value if not isinstance(value, list) else ", ".join(str(v) for v in value)
                )
        return flat

    def add_sheet(sheet_name: str, records: list[dict]) -> None:
        """Add a sheet with records."""
        if not records:
            return

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        # Collect all unique keys preserving order
        all_keys = []
        for record in records:
            for key in record:
                if key not in all_keys:
                    all_keys.append(key)

        # Write header
        for col, key in enumerate(all_keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row, record in enumerate(records, 2):
            for col, key in enumerate(all_keys, 1):
                value = record.get(key, "")
                if isinstance(value, list | dict):
                    value = str(value)
                ws.cell(row=row, column=col, value=value)

        # Auto-width columns
        for col in range(1, len(all_keys) + 1):
            max_length = max(
                len(str(ws.cell(row=row, column=col).value or ""))
                for row in range(1, len(records) + 2)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

    # Root entity (Investigation)
    root_record = flatten_entity(data)
    add_sheet("Investigation", [root_record])

    # Contacts/Persons at investigation level
    if data.get("contacts"):
        records = [flatten_entity(c) for c in data["contacts"]]
        add_sheet("Person", records)

    # Studies
    if data.get("studies"):
        study_records = []
        all_persons = []
        all_locations = []
        all_bio_materials = []
        all_factors = []
        all_factor_values = []
        all_obs_variables = []
        all_obs_units = []
        all_samples = []
        all_events = []
        all_environments = []
        all_data_files = []
        all_protocols = []
        all_sources = []
        all_assays = []

        for study in data["studies"]:
            study_records.append(flatten_entity(study))

            # Nested entities within study
            if "persons" in study:
                all_persons.extend(flatten_entity(p) for p in study["persons"])
            if study.get("geographic_location"):
                loc = study["geographic_location"]
                if isinstance(loc, dict):
                    all_locations.append(flatten_entity(loc))
            if "biological_materials" in study:
                for bm in study["biological_materials"]:
                    all_bio_materials.append(flatten_entity(bm))
            if "factors" in study:
                for f in study["factors"]:
                    all_factors.append(flatten_entity(f))
                    if "values" in f:
                        all_factor_values.extend(flatten_entity(fv) for fv in f["values"])
            if "observed_variables" in study:
                all_obs_variables.extend(flatten_entity(ov) for ov in study["observed_variables"])
            if "observation_units" in study:
                for ou in study["observation_units"]:
                    all_obs_units.append(flatten_entity(ou))
                    if "samples" in ou:
                        all_samples.extend(flatten_entity(s) for s in ou["samples"])
                    if "factor_values" in ou:
                        all_factor_values.extend(flatten_entity(fv) for fv in ou["factor_values"])
            if "events" in study:
                all_events.extend(flatten_entity(e) for e in study["events"])
            if "environments" in study:
                all_environments.extend(flatten_entity(env) for env in study["environments"])
            if "data_files" in study:
                all_data_files.extend(flatten_entity(df) for df in study["data_files"])
            if "protocols" in study:
                all_protocols.extend(flatten_entity(p) for p in study["protocols"])
            if "sources" in study:
                all_sources.extend(flatten_entity(s) for s in study["sources"])
            if "samples" in study:
                all_samples.extend(flatten_entity(s) for s in study["samples"])
            if "assays" in study:
                all_assays.extend(flatten_entity(a) for a in study["assays"])

        add_sheet("Study", study_records)
        add_sheet("Person", all_persons) if all_persons else None
        add_sheet("Location", all_locations) if all_locations else None
        add_sheet("BiologicalMaterial", all_bio_materials) if all_bio_materials else None
        add_sheet("Factor", all_factors) if all_factors else None
        add_sheet("FactorValue", all_factor_values) if all_factor_values else None
        add_sheet("ObservedVariable", all_obs_variables) if all_obs_variables else None
        add_sheet("ObservationUnit", all_obs_units) if all_obs_units else None
        add_sheet("Sample", all_samples) if all_samples else None
        add_sheet("Event", all_events) if all_events else None
        add_sheet("Environment", all_environments) if all_environments else None
        add_sheet("DataFile", all_data_files) if all_data_files else None
        add_sheet("Protocol", all_protocols) if all_protocols else None
        add_sheet("Source", all_sources) if all_sources else None
        add_sheet("Assay", all_assays) if all_assays else None

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def _parse_profile_spec(spec: str) -> tuple[str, str]:
    """Parse a profile/version specification.

    Args:
        spec: Profile spec in format "profile/version" (e.g., "miappe/1.1").

    Returns:
        Tuple of (profile, version).

    Raises:
        typer.Exit: If format is invalid.
    """
    if "/" not in spec:
        echo_error(f"Invalid profile format: '{spec}'. Use 'profile/version' (e.g., 'miappe/1.1')")
        raise typer.Exit(EXIT_INPUT_ERROR)

    parts = spec.split("/", 1)
    return parts[0], parts[1]


@app.command(name="compare")
def compare_profiles(
    profiles: Annotated[
        list[str], typer.Argument(help="Profiles to compare (format: profile/version)")
    ],
    output: Annotated[Path | None, typer.Option("--output", "-o", help="Output file path")] = None,
    format: Annotated[
        str, typer.Option("--format", "-f", help="Output format: markdown, csv, html")
    ] = "markdown",
) -> None:
    """Compare multiple profile specifications.

    Shows differences in entities, fields, and constraints across profiles.

    Examples:
        metaseed compare miappe/1.1 isa/1.0
        metaseed compare miappe/1.1 miappe/1.2 -o diff.md -f markdown
        metaseed compare miappe/1.1 isa/1.0 cropxr-phenotyping/0.0.5
    """
    if len(profiles) < 2:
        echo_error("At least 2 profiles required for comparison")
        raise typer.Exit(EXIT_INPUT_ERROR)

    # Parse profile specs
    profile_tuples = []
    for spec in profiles:
        profile_tuples.append(_parse_profile_spec(spec))

    try:
        from metaseed.specs.merge import (
            CSVReportGenerator,
            HTMLReportGenerator,
            MarkdownReportGenerator,
            compare,
        )

        result = compare(profile_tuples)

        # Generate report
        if format.lower() == "csv":
            report = CSVReportGenerator(result).generate()
        elif format.lower() == "html":
            report = HTMLReportGenerator(result).generate()
        else:
            report = MarkdownReportGenerator(result).generate()

        if output:
            output.parent.mkdir(parents=True, exist_ok=True)
            output.write_text(report, encoding="utf-8")
            echo_success(f"Comparison report written to {output}")
        else:
            typer.echo(report)

        # Print summary
        stats = result.statistics
        typer.echo("")
        typer.echo(f"Compared {len(profiles)} profiles:")
        typer.echo(f"  Entities: {stats.total_entities} total, {stats.common_entities} common")
        typer.echo(f"  Fields: {stats.total_fields} total, {stats.conflicting_fields} conflicts")

    except SpecLoadError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_CONFIG_ERROR) from None
    except ValueError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_INPUT_ERROR) from None


@app.command(name="merge")
def merge_profiles(
    profiles: Annotated[
        list[str], typer.Argument(help="Profiles to merge (format: profile/version)")
    ],
    output: Annotated[Path, typer.Option("--output", "-o", help="Output YAML file path")] = Path(
        "merged.yaml"
    ),
    strategy: Annotated[
        str, typer.Option("--strategy", "-s", help="Merge strategy")
    ] = "first_wins",
    name: Annotated[
        str | None, typer.Option("--name", "-n", help="Name for merged profile")
    ] = None,
    version: Annotated[
        str, typer.Option("--version", "-v", help="Version for merged profile")
    ] = "1.0",
) -> None:
    """Merge multiple profile specifications into one.

    Combines entities and fields from multiple profiles with conflict resolution.

    Strategies:
        first_wins: Use first profile's values for conflicts
        last_wins: Use last profile's values for conflicts
        most_restrictive: required=True wins, tighter constraints
        least_restrictive: required=False wins, looser constraints
        prefer_<profile>: Always prefer specific profile (e.g., prefer_miappe/1.1)

    Examples:
        metaseed merge miappe/1.1 isa/1.0 -o combined.yaml
        metaseed merge miappe/1.1 cropxr-phenotyping/0.0.5 -s most_restrictive -o strict.yaml
        metaseed merge miappe/1.1 isa/1.0 -s prefer_miappe/1.1 -n miappe-extended
    """
    if len(profiles) < 2:
        echo_error("At least 2 profiles required for merge")
        raise typer.Exit(EXIT_INPUT_ERROR)

    # Parse profile specs
    profile_tuples = []
    for spec in profiles:
        profile_tuples.append(_parse_profile_spec(spec))

    # Default name from profiles
    if name is None:
        name = "-".join(p[0] for p in profile_tuples) + "-merged"

    try:
        from metaseed.specs.merge import merge

        result = merge(
            profiles=profile_tuples,
            strategy=strategy,
            output_name=name,
            output_version=version,
        )

        # Write output
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(result.to_yaml(), encoding="utf-8")

        echo_success(f"Merged profile written to {output}")

        # Print summary
        typer.echo(f"  Name: {result.merged_profile.name}")
        typer.echo(f"  Version: {result.merged_profile.version}")
        typer.echo(f"  Entities: {len(result.merged_profile.entities)}")
        typer.echo(f"  Strategy: {result.strategy_used}")

        if result.warnings:
            typer.echo(f"  Warnings: {len(result.warnings)}")

        if result.has_unresolved_conflicts:
            typer.echo(f"  Unresolved conflicts: {len(result.unresolved_conflicts)}")

    except SpecLoadError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_CONFIG_ERROR) from None
    except ValueError as e:
        echo_error(str(e))
        raise typer.Exit(EXIT_INPUT_ERROR) from None


if __name__ == "__main__":
    app()
