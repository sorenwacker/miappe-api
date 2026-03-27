"""HTMX route handlers for the UI.

Provides FastAPI routes with Jinja2 templates for the HTMX-based interface.
"""

from __future__ import annotations

import contextlib
import uuid
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import ValidationError

from miappe_api.facade import ProfileFacade
from miappe_api.validators import validate as validate_data

UI_DIR = Path(__file__).parent
TEMPLATES_DIR = UI_DIR / "templates"
STATIC_DIR = UI_DIR / "static"


@dataclass
class TreeNode:
    """A node in the entity tree."""

    id: str
    entity_type: str
    instance: Any
    label: str
    children: list[TreeNode] = field(default_factory=list)
    parent_id: str | None = None

    @classmethod
    def create(cls, entity_type: str, instance: Any, parent_id: str | None = None) -> TreeNode:
        """Create a new tree node from an entity instance."""
        label = ""
        if hasattr(instance, "model_dump"):
            data = instance.model_dump()
            for key in ["title", "name", "unique_id", "identifier", "filename"]:
                if key in data and data[key]:
                    label = str(data[key])
                    break
        if not label:
            label = f"New {entity_type}"

        return cls(
            id=str(uuid.uuid4())[:8],
            entity_type=entity_type,
            instance=instance,
            label=label,
            parent_id=parent_id,
        )

    def to_dict(self) -> dict:
        """Convert node to dictionary for template rendering."""
        return {
            "id": self.id,
            "entity_type": self.entity_type,
            "label": self.label,
            "has_children": bool(self.children),
            "children": [c.to_dict() for c in self.children],
        }


@dataclass
class NestedEditContext:
    """Context for editing a nested item."""

    field_name: str  # The field containing this item (e.g., "studies")
    row_idx: int  # Index in the parent's list
    entity_type: str  # Entity type being edited (e.g., "Study")
    parent_entity_type: str  # Parent entity type (e.g., "Investigation")
    nested_items: dict[str, list] = field(default_factory=dict)  # This item's nested fields


@dataclass
class AppState:
    """Server-side state for the UI."""

    profile: str = "miappe"
    facade: ProfileFacade | None = None
    entity_tree: list[TreeNode] = field(default_factory=list)
    nodes_by_id: dict[str, TreeNode] = field(default_factory=dict)
    editing_node_id: str | None = None
    current_nested_items: dict[str, list] = field(default_factory=dict)
    nested_edit_stack: list[NestedEditContext] = field(default_factory=list)  # Navigation stack

    def get_or_create_facade(self) -> ProfileFacade:
        """Get existing facade or create new one."""
        if self.facade is None or self.facade.profile != self.profile:
            self.facade = ProfileFacade(self.profile)
        return self.facade

    def get_root_entity_types(self) -> list[str]:
        """Get entity types that can be created at root level."""
        facade = self.get_or_create_facade()
        referenced_by: dict[str, set[str]] = {name: set() for name in facade.entities}

        for entity_name in facade.entities:
            helper = getattr(facade, entity_name)
            for ref_entity in helper.nested_fields.values():
                if ref_entity in referenced_by:
                    referenced_by[ref_entity].add(entity_name)

        roots = []
        for name in facade.entities:
            refs = referenced_by[name] - {name}
            if not refs:
                roots.append(name)

        if "Investigation" in roots:
            roots.remove("Investigation")
            roots.insert(0, "Investigation")

        return roots

    def add_node(self, entity_type: str, instance: Any, parent_id: str | None = None) -> TreeNode:
        """Add a new node to the tree."""
        node = TreeNode.create(entity_type, instance, parent_id)
        self.nodes_by_id[node.id] = node

        if parent_id and parent_id in self.nodes_by_id:
            self.nodes_by_id[parent_id].children.append(node)
        else:
            self.entity_tree.append(node)

        return node

    def update_node(self, node_id: str, instance: Any) -> TreeNode | None:
        """Update an existing node."""
        node = self.nodes_by_id.get(node_id)
        if node:
            node.instance = instance
            if hasattr(instance, "model_dump"):
                data = instance.model_dump()
                for key in ["title", "name", "unique_id", "identifier", "filename"]:
                    if key in data and data[key]:
                        node.label = str(data[key])
                        break
        return node

    def delete_node(self, node_id: str) -> bool:
        """Delete a node and all its children."""
        node = self.nodes_by_id.get(node_id)
        if not node:
            return False

        def remove_recursively(n: TreeNode) -> None:
            for child in n.children:
                remove_recursively(child)
            self.nodes_by_id.pop(n.id, None)

        if node.parent_id and node.parent_id in self.nodes_by_id:
            parent = self.nodes_by_id[node.parent_id]
            parent.children = [c for c in parent.children if c.id != node.id]
        else:
            self.entity_tree = [n for n in self.entity_tree if n.id != node.id]

        remove_recursively(node)

        if self.editing_node_id == node_id:
            self.editing_node_id = None

        return True

    def get_tree_data(self) -> list[dict]:
        """Get tree data for template rendering, including nested entities."""
        facade = self.get_or_create_facade()

        def extract_nested_children(data: dict, entity_type: str) -> list[dict]:
            """Extract nested entities as tree children."""
            children = []
            helper = getattr(facade, entity_type, None)
            if not helper:
                return children

            for field_name, nested_type in helper.nested_fields.items():
                nested_items = data.get(field_name, [])
                if not nested_items or not isinstance(nested_items, list):
                    continue

                for i, item in enumerate(nested_items):
                    if hasattr(item, "model_dump"):
                        item_data = item.model_dump(exclude_none=True)
                    elif isinstance(item, dict):
                        item_data = item
                    else:
                        continue

                    # Get label for nested item
                    nested_helper = getattr(facade, nested_type, None)
                    label = None
                    if nested_helper:
                        for field in ["title", "name", "unique_id", "identifier"]:
                            if field in item_data and item_data[field]:
                                label = str(item_data[field])
                                break
                    if not label:
                        label = f"{nested_type} {i + 1}"

                    child = {
                        "id": f"{field_name}_{i}",
                        "entity_type": nested_type,
                        "label": label,
                        "field_name": field_name,
                        "idx": i,
                        "parent_entity_type": entity_type,
                        "is_nested": True,
                        "has_children": False,
                        "children": [],
                    }

                    # Recursively get nested children
                    nested_children = extract_nested_children(item_data, nested_type)
                    if nested_children:
                        child["has_children"] = True
                        child["children"] = nested_children

                    children.append(child)

            return children

        def node_to_dict_with_nested(node: TreeNode) -> dict:
            """Convert node to dict including nested entities as children."""
            result = {
                "id": node.id,
                "entity_type": node.entity_type,
                "label": node.label,
                "has_children": False,
                "children": [],
            }

            # Get nested children from instance data
            if node.instance and hasattr(node.instance, "model_dump"):
                data = node.instance.model_dump(exclude_none=True)
                nested_children = extract_nested_children(data, node.entity_type)
                if nested_children:
                    result["has_children"] = True
                    result["children"] = nested_children

            return result

        return [node_to_dict_with_nested(n) for n in self.entity_tree]


def create_app(state: AppState | None = None) -> FastAPI:
    """Create the FastAPI application with HTMX routes.

    Args:
        state: Optional initial state. Creates new state if not provided.

    Returns:
        Configured FastAPI application.
    """
    app = FastAPI(title="MIAPPE-API UI")

    if state is None:
        state = AppState()

    app.state.ui_state = state

    templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

    def get_state() -> AppState:
        return app.state.ui_state

    @app.get("/", response_class=HTMLResponse)
    async def index(request: Request):
        """Render the main page."""
        state = get_state()
        facade = state.get_or_create_facade()

        return templates.TemplateResponse(
            request,
            "base.html",
            {
                "profiles": ["miappe", "isa", "combined"],
                "current_profile": state.profile,
                "version": facade.version,
                "root_types": state.get_root_entity_types()[:3],
                "tree_nodes": state.get_tree_data(),
                "editing_node_id": state.editing_node_id,
            },
        )

    @app.get("/form/{entity_type}", response_class=HTMLResponse)
    async def new_entity_form(request: Request, entity_type: str):
        """Render a new entity form."""
        state = get_state()
        facade = state.get_or_create_facade()

        try:
            helper = getattr(facade, entity_type)
        except AttributeError as e:
            raise HTTPException(
                status_code=404, detail=f"Entity type not found: {entity_type}"
            ) from e

        state.editing_node_id = None
        state.current_nested_items = {}  # Reset nested items for new entity

        fields = _get_field_data(helper)

        # Auto-populate values for certain fields
        auto_values = {}
        if "miappe_version" in helper.all_fields:
            auto_values["miappe_version"] = facade.version

        return templates.TemplateResponse(
            request,
            "partials/form.html",
            {
                "entity_type": entity_type,
                "is_edit": False,
                "node_id": None,
                "description": helper.description,
                "ontology_term": helper.ontology_term,
                "required_fields": [f for f in fields if f["required"]],
                "optional_fields": [
                    f for f in fields if not f["required"] and not _is_nested_field(f)
                ],
                "nested_fields": [f for f in fields if _is_nested_field(f)],
                "values": auto_values,
                "auto_fields": set(auto_values.keys()),
            },
        )

    @app.get("/form/{entity_type}/{node_id}", response_class=HTMLResponse)
    async def edit_entity_form(request: Request, entity_type: str, node_id: str):
        """Render an edit form for an existing entity."""
        state = get_state()
        facade = state.get_or_create_facade()

        node = state.nodes_by_id.get(node_id)
        if not node:
            raise HTTPException(status_code=404, detail=f"Node not found: {node_id}")

        try:
            helper = getattr(facade, entity_type)
        except AttributeError as e:
            raise HTTPException(
                status_code=404, detail=f"Entity type not found: {entity_type}"
            ) from e

        state.editing_node_id = node_id

        fields = _get_field_data(helper)
        values = {}
        if node.instance and hasattr(node.instance, "model_dump"):
            values = node.instance.model_dump(exclude_none=True)

        # Merge current_nested_items (from table editing) into values for display
        # This preserves data from table edits that haven't been saved yet
        for field_name, items in state.current_nested_items.items():
            if items:
                values[field_name] = items

        # Only load nested items from entity if we don't have any pending edits
        if not state.current_nested_items:
            # Load nested items from entity for table editing
            for field_name in helper.nested_fields:
                if field_name in values and values[field_name]:
                    items = values[field_name]
                    if isinstance(items, list):
                        state.current_nested_items[field_name] = [
                            item.model_dump() if hasattr(item, "model_dump") else item
                            for item in items
                        ]

        # Auto-populate values for certain fields
        auto_fields = set()
        if "miappe_version" in helper.all_fields:
            values["miappe_version"] = facade.version
            auto_fields.add("miappe_version")

        return templates.TemplateResponse(
            request,
            "partials/form.html",
            {
                "entity_type": entity_type,
                "is_edit": True,
                "node_id": node_id,
                "description": helper.description,
                "ontology_term": helper.ontology_term,
                "required_fields": [f for f in fields if f["required"]],
                "optional_fields": [
                    f for f in fields if not f["required"] and not _is_nested_field(f)
                ],
                "nested_fields": [f for f in fields if _is_nested_field(f)],
                "values": values,
                "auto_fields": auto_fields,
            },
        )

    @app.post("/entity", response_class=HTMLResponse)
    async def create_entity(request: Request):
        """Create a new entity."""
        state = get_state()
        facade = state.get_or_create_facade()

        form_data = await request.form()
        entity_type = form_data.get("_entity_type")

        if not entity_type:
            return _error_response(request, templates, "Entity type is required")

        try:
            helper = getattr(facade, entity_type)
        except AttributeError:
            return _error_response(request, templates, f"Unknown entity type: {entity_type}")

        values = _collect_form_values(form_data, helper)

        try:
            instance = helper.create(**values)
            node = state.add_node(entity_type, instance)
            state.editing_node_id = node.id

            # Load nested items from the new entity for editing
            state.current_nested_items = {}
            if hasattr(instance, "model_dump"):
                data = instance.model_dump(exclude_none=True)
                for field_name in helper.nested_fields:
                    if field_name in data and data[field_name]:
                        items = data[field_name]
                        if isinstance(items, list):
                            state.current_nested_items[field_name] = [
                                item.model_dump() if hasattr(item, "model_dump") else item
                                for item in items
                            ]

            # Return the edit form for the newly created entity
            fields = _get_field_data(helper)
            edit_values = (
                instance.model_dump(exclude_none=True) if hasattr(instance, "model_dump") else {}
            )

            auto_fields = set()
            if "miappe_version" in helper.all_fields:
                edit_values["miappe_version"] = facade.version
                auto_fields.add("miappe_version")

            response = templates.TemplateResponse(
                request,
                "partials/form.html",
                {
                    "entity_type": entity_type,
                    "is_edit": True,
                    "node_id": node.id,
                    "description": helper.description,
                    "ontology_term": helper.ontology_term,
                    "required_fields": [f for f in fields if f["required"]],
                    "optional_fields": [
                        f for f in fields if not f["required"] and not _is_nested_field(f)
                    ],
                    "nested_fields": [f for f in fields if _is_nested_field(f)],
                    "values": edit_values,
                    "auto_fields": auto_fields,
                    "success_message": f"Created {entity_type}: {node.label}",
                },
            )
            response.headers["HX-Trigger"] = "entityCreated"
            return response

        except ValidationError as e:
            errors = _format_validation_errors(e)
            # Re-render form with errors and user's data preserved
            fields = _get_field_data(helper)
            auto_values = {}
            if "miappe_version" in helper.all_fields:
                auto_values["miappe_version"] = facade.version
            values.update(auto_values)

            return templates.TemplateResponse(
                request,
                "partials/form.html",
                {
                    "entity_type": entity_type,
                    "is_edit": False,
                    "node_id": None,
                    "description": helper.description,
                    "ontology_term": helper.ontology_term,
                    "required_fields": [f for f in fields if f["required"]],
                    "optional_fields": [
                        f for f in fields if not f["required"] and not _is_nested_field(f)
                    ],
                    "nested_fields": [f for f in fields if _is_nested_field(f)],
                    "values": values,
                    "auto_fields": set(auto_values.keys()),
                    "error_message": f"Validation error: {errors}",
                },
            )

    @app.put("/entity/{node_id}", response_class=HTMLResponse)
    async def update_entity(request: Request, node_id: str):
        """Update an existing entity."""
        state = get_state()
        facade = state.get_or_create_facade()

        node = state.nodes_by_id.get(node_id)
        if not node:
            return _error_response(request, templates, f"Node not found: {node_id}")

        form_data = await request.form()
        entity_type = node.entity_type

        try:
            helper = getattr(facade, entity_type)
        except AttributeError:
            return _error_response(request, templates, f"Unknown entity type: {entity_type}")

        values = _collect_form_values(form_data, helper)

        # Merge nested items from table editing
        for field_name, items in state.current_nested_items.items():
            if field_name in helper.nested_fields and items:
                cleaned_items = []
                for item in items:
                    if isinstance(item, dict):
                        # Remove internal keys and empty strings
                        cleaned = {k: v for k, v in item.items() if not k.startswith("_") and v}
                        if cleaned:
                            cleaned_items.append(cleaned)
                if cleaned_items:
                    values[field_name] = cleaned_items

        try:
            instance = helper.create(**values)
            state.update_node(node_id, instance)

            # Reload nested items from saved instance
            state.current_nested_items = {}
            if hasattr(instance, "model_dump"):
                data = instance.model_dump(exclude_none=True)
                for field_name in helper.nested_fields:
                    if field_name in data and data[field_name]:
                        items = data[field_name]
                        if isinstance(items, list):
                            state.current_nested_items[field_name] = [
                                item.model_dump() if hasattr(item, "model_dump") else item
                                for item in items
                            ]

            # Stay on the form with success message
            fields = _get_field_data(helper)
            edit_values = (
                instance.model_dump(exclude_none=True) if hasattr(instance, "model_dump") else {}
            )

            auto_fields = set()
            if "miappe_version" in helper.all_fields:
                edit_values["miappe_version"] = facade.version
                auto_fields.add("miappe_version")

            response = templates.TemplateResponse(
                request,
                "partials/form.html",
                {
                    "entity_type": entity_type,
                    "is_edit": True,
                    "node_id": node_id,
                    "description": helper.description,
                    "ontology_term": helper.ontology_term,
                    "required_fields": [f for f in fields if f["required"]],
                    "optional_fields": [
                        f for f in fields if not f["required"] and not _is_nested_field(f)
                    ],
                    "nested_fields": [f for f in fields if _is_nested_field(f)],
                    "values": edit_values,
                    "auto_fields": auto_fields,
                    "success_message": f"Saved {entity_type}: {node.label}",
                },
            )
            response.headers["HX-Trigger"] = "entityUpdated"
            return response

        except ValidationError as e:
            errors = _format_validation_errors(e)
            # Re-render form with errors and user's data preserved
            fields = _get_field_data(helper)
            auto_fields = set()
            if "miappe_version" in helper.all_fields:
                values["miappe_version"] = facade.version
                auto_fields.add("miappe_version")

            return templates.TemplateResponse(
                request,
                "partials/form.html",
                {
                    "entity_type": entity_type,
                    "is_edit": True,
                    "node_id": node_id,
                    "description": helper.description,
                    "ontology_term": helper.ontology_term,
                    "required_fields": [f for f in fields if f["required"]],
                    "optional_fields": [
                        f for f in fields if not f["required"] and not _is_nested_field(f)
                    ],
                    "nested_fields": [f for f in fields if _is_nested_field(f)],
                    "values": values,
                    "auto_fields": auto_fields,
                    "error_message": f"Validation error: {errors}",
                },
            )

    @app.delete("/entity/{node_id}", response_class=HTMLResponse)
    async def delete_entity(request: Request, node_id: str):
        """Delete an entity."""
        state = get_state()

        node = state.nodes_by_id.get(node_id)
        if not node:
            return _error_response(request, templates, f"Node not found: {node_id}")

        entity_type = node.entity_type
        label = node.label

        state.delete_node(node_id)

        # Return welcome page with delete notification
        return templates.TemplateResponse(
            request,
            "index.html",
            {
                "notification": {
                    "type": "warning",
                    "message": f"Deleted {entity_type}: {label}",
                },
            },
        )

    @app.get("/table/{entity_type}/{field_name}", response_class=HTMLResponse)
    async def table_view(request: Request, entity_type: str, field_name: str):
        """Render the nested table view for a list field."""
        state = get_state()
        facade = state.get_or_create_facade()

        try:
            helper = getattr(facade, entity_type)
        except AttributeError as e:
            raise HTTPException(
                status_code=404, detail=f"Entity type not found: {entity_type}"
            ) from e

        nested_fields = helper.nested_fields
        if field_name not in nested_fields:
            raise HTTPException(status_code=404, detail=f"Field not found: {field_name}")

        nested_entity_type = nested_fields[field_name]
        nested_helper = getattr(facade, nested_entity_type, None)

        columns = []
        column_types = {}
        column_patterns = {}
        required_columns = set()
        has_nested_children = False
        if nested_helper:
            cols = list(nested_helper.required_fields) + list(nested_helper.optional_fields)
            nested = set(nested_helper.nested_fields.keys())
            columns = [c for c in cols if c not in nested]
            required_columns = set(nested_helper.required_fields)
            has_nested_children = bool(nested_helper.nested_fields)
            for col in columns:
                info = nested_helper.field_info(col)
                column_types[col] = info.get("type", "string")
                # Extract pattern from constraints if available
                constraints = info.get("constraints", {})
                if constraints and "pattern" in constraints:
                    column_patterns[col] = constraints["pattern"]
        else:
            columns = ["value"]
            column_types["value"] = "string"

        # Check if we're in a nested context and use context's items if appropriate
        nested_context = state.nested_edit_stack[-1] if state.nested_edit_stack else None
        if nested_context and field_name in nested_context.nested_items:
            items = nested_context.nested_items.get(field_name, [])
        else:
            items = state.current_nested_items.get(field_name, [])

        rows = []
        for i, item in enumerate(items):
            if hasattr(item, "model_dump"):
                row = item.model_dump(exclude_none=True)
            elif isinstance(item, dict):
                row = item.copy()
            else:
                row = {"value": str(item)}
            row["_idx"] = i
            rows.append(row)

        # Determine if we're in a nested context (editing nested item)
        nested_context = state.nested_edit_stack[-1] if state.nested_edit_stack else None

        return templates.TemplateResponse(
            request,
            "partials/table.html",
            {
                "field_name": field_name,
                "entity_type": nested_entity_type,
                "columns": columns,
                "column_types": column_types,
                "column_patterns": column_patterns,
                "required_columns": required_columns,
                "has_nested_children": has_nested_children,
                "rows": rows,
                "parent_entity_type": entity_type,
                "editing_node_id": state.editing_node_id,
                "breadcrumb": _build_breadcrumb(state),
                "nested_context": nested_context,
            },
        )

    @app.post("/table/{parent_entity_type}/{field_name}/row", response_class=HTMLResponse)
    async def add_table_row(request: Request, parent_entity_type: str, field_name: str):
        """Add a new row to the nested table."""
        state = get_state()
        facade = state.get_or_create_facade()

        # Determine where to store items (context or root state)
        nested_context = state.nested_edit_stack[-1] if state.nested_edit_stack else None
        if nested_context and field_name in nested_context.nested_items:
            items_store = nested_context.nested_items
        else:
            items_store = state.current_nested_items

        if field_name not in items_store:
            items_store[field_name] = []

        form_data = await request.form()
        # Extract column names from _col_* hidden inputs
        columns = [form_data[k] for k in form_data if k.startswith("_col_")]

        new_row = dict.fromkeys(columns, "")
        new_row["_idx"] = len(items_store[field_name])
        items_store[field_name].append(new_row)

        # Get entity type for the row and column types
        parent_helper = getattr(facade, parent_entity_type, None)
        entity_type = parent_helper.nested_fields.get(field_name) if parent_helper else None

        # Get column types and patterns for the nested entity
        column_types = {}
        column_patterns = {}
        if entity_type:
            nested_helper = getattr(facade, entity_type, None)
            if nested_helper:
                for col in columns:
                    info = nested_helper.field_info(col)
                    column_types[col] = info.get("type", "string")
                    constraints = info.get("constraints", {})
                    if constraints and "pattern" in constraints:
                        column_patterns[col] = constraints["pattern"]

        return templates.TemplateResponse(
            request,
            "partials/table_row.html",
            {
                "row": new_row,
                "columns": columns,
                "column_types": column_types,
                "column_patterns": column_patterns,
                "field_name": field_name,
                "parent_entity_type": parent_entity_type,
                "entity_type": entity_type,
            },
        )

    @app.delete("/table/{field_name}/row/{idx}", response_class=HTMLResponse)
    async def delete_table_row(field_name: str, idx: int):
        """Delete a row from the nested table."""
        state = get_state()

        # Determine which items store to use
        nested_context = state.nested_edit_stack[-1] if state.nested_edit_stack else None
        if nested_context and field_name in nested_context.nested_items:
            items_store = nested_context.nested_items
        else:
            items_store = state.current_nested_items

        if field_name in items_store:
            items = items_store[field_name]
            if 0 <= idx < len(items):
                del items[idx]
                for i, item in enumerate(items):
                    if isinstance(item, dict):
                        item["_idx"] = i

        return HTMLResponse(content="")

    @app.post("/table/{field_name}/row/{idx}/cell", response_class=HTMLResponse)
    async def update_table_cell(request: Request, field_name: str, idx: int):
        """Update a cell value in the nested table."""
        state = get_state()

        # Determine which items store to use
        nested_context = state.nested_edit_stack[-1] if state.nested_edit_stack else None
        if nested_context and field_name in nested_context.nested_items:
            items_store = nested_context.nested_items
        else:
            items_store = state.current_nested_items

        if field_name in items_store:
            items = items_store[field_name]
            if 0 <= idx < len(items):
                form_data = await request.form()
                item = items[idx]
                if isinstance(item, dict):
                    for key, value in form_data.items():
                        if not key.startswith("_"):
                            item[key] = value

        return HTMLResponse(content="")

    @app.get("/nested/{parent_type}/{field_name}/{idx}", response_class=HTMLResponse)
    async def edit_nested_item(request: Request, parent_type: str, field_name: str, idx: int):
        """Edit a nested item (e.g., a Study within an Investigation).

        Query params:
            resume: If true, don't push a new context (returning from child table)
        """
        state = get_state()
        facade = state.get_or_create_facade()
        is_resume = request.query_params.get("resume") == "true"

        # Get parent entity helper to find nested entity type
        try:
            parent_helper = getattr(facade, parent_type)
        except AttributeError as e:
            raise HTTPException(
                status_code=404, detail=f"Entity type not found: {parent_type}"
            ) from e

        if field_name not in parent_helper.nested_fields:
            raise HTTPException(status_code=404, detail=f"Field not found: {field_name}")

        nested_entity_type = parent_helper.nested_fields[field_name]
        nested_helper = getattr(facade, nested_entity_type, None)

        # For resume: use existing context's nested items
        # For fresh edit: get items from parent's current_nested_items
        if is_resume and state.nested_edit_stack:
            # Use the top of stack context's nested items
            context = state.nested_edit_stack[-1]
            # Get item data from parent items, merge with context nested items
            parent_items = state.current_nested_items.get(field_name, [])
            if idx < len(parent_items):
                item_data = parent_items[idx]
                if hasattr(item_data, "model_dump"):
                    item_data = item_data.model_dump(exclude_none=True)
                elif isinstance(item_data, dict):
                    item_data = item_data.copy()
                else:
                    item_data = {}
                # Merge context nested items back
                for nf, nv in context.nested_items.items():
                    item_data[nf] = nv
            else:
                item_data = {}
        else:
            # Get the nested item data from parent's nested items
            items = state.current_nested_items.get(field_name, [])
            if idx < 0 or idx >= len(items):
                raise HTTPException(status_code=404, detail=f"Row not found: {idx}")

            item_data = items[idx]
            if hasattr(item_data, "model_dump"):
                item_data = item_data.model_dump(exclude_none=True)
            elif isinstance(item_data, dict):
                item_data = item_data.copy()
            else:
                item_data = {}

            # Push context onto stack
            context = NestedEditContext(
                field_name=field_name,
                row_idx=idx,
                entity_type=nested_entity_type,
                parent_entity_type=parent_type,
            )

            # Load nested items for this item
            if nested_helper:
                for nested_field in nested_helper.nested_fields:
                    if nested_field in item_data and item_data[nested_field]:
                        nested_items = item_data[nested_field]
                        if isinstance(nested_items, list):
                            context.nested_items[nested_field] = [
                                i.model_dump() if hasattr(i, "model_dump") else i
                                for i in nested_items
                            ]

            state.nested_edit_stack.append(context)

        # Get field info for the nested entity
        fields = _get_field_data(nested_helper) if nested_helper else []

        # Merge context nested items into values for counter display
        values = item_data.copy() if isinstance(item_data, dict) else {}
        if state.nested_edit_stack:
            ctx = state.nested_edit_stack[-1]
            for nf, nv in ctx.nested_items.items():
                values[nf] = nv

        return templates.TemplateResponse(
            request,
            "partials/nested_form.html",
            {
                "entity_type": nested_entity_type,
                "parent_entity_type": parent_type,
                "field_name": field_name,
                "row_idx": idx,
                "description": nested_helper.description if nested_helper else "",
                "ontology_term": nested_helper.ontology_term if nested_helper else "",
                "required_fields": [f for f in fields if f["required"]],
                "optional_fields": [
                    f for f in fields if not f["required"] and not _is_nested_field(f)
                ],
                "nested_fields": [f for f in fields if _is_nested_field(f)],
                "values": values,
                "editing_node_id": state.editing_node_id,
                "breadcrumb": _build_breadcrumb(state),
            },
        )

    @app.post("/nested/{parent_type}/{field_name}/{idx}", response_class=HTMLResponse)
    async def save_nested_item(request: Request, parent_type: str, field_name: str, idx: int):
        """Save changes to a nested item and return to parent table."""
        state = get_state()
        facade = state.get_or_create_facade()

        # Get the nested item
        items = state.current_nested_items.get(field_name, [])
        if idx < 0 or idx >= len(items):
            raise HTTPException(status_code=404, detail=f"Row not found: {idx}")

        # Update item data from form
        form_data = await request.form()

        # Get nested entity helper for field info
        parent_helper = getattr(facade, parent_type, None)
        nested_entity_type = parent_helper.nested_fields.get(field_name) if parent_helper else None
        nested_helper = getattr(facade, nested_entity_type, None) if nested_entity_type else None

        item = items[idx]
        if isinstance(item, dict):
            for key, value in form_data.items():
                if not key.startswith("_"):
                    # Convert value types if needed
                    if nested_helper:
                        info = (
                            nested_helper.field_info(key) if key in nested_helper.all_fields else {}
                        )
                        if info.get("type") == "integer" and value:
                            with contextlib.suppress(ValueError):
                                value = int(value)
                        elif info.get("type") == "float" and value:
                            with contextlib.suppress(ValueError):
                                value = float(value)
                    if value:  # Only set non-empty values
                        item[key] = value

            # Merge nested items from stack context
            if state.nested_edit_stack:
                context = state.nested_edit_stack[-1]
                for nested_field, nested_values in context.nested_items.items():
                    if nested_values:
                        item[nested_field] = nested_values

        # Pop from stack
        if state.nested_edit_stack:
            state.nested_edit_stack.pop()

        # Return to parent table
        return templates.TemplateResponse(
            request,
            "partials/table.html",
            {
                "field_name": field_name,
                "entity_type": nested_entity_type,
                "columns": _get_table_columns(facade, nested_entity_type),
                "rows": _format_table_rows(items),
                "parent_entity_type": parent_type,
                "editing_node_id": state.editing_node_id,
            },
        )

    @app.get("/profile/{name}")
    async def switch_profile(name: str):
        """Switch to a different profile."""
        state = get_state()

        if name not in ["miappe", "isa", "combined"]:
            raise HTTPException(status_code=400, detail=f"Unknown profile: {name}")

        state.profile = name
        state.facade = None
        state.entity_tree = []
        state.nodes_by_id = {}
        state.editing_node_id = None
        state.current_nested_items = {}
        state.nested_edit_stack = []

        return RedirectResponse(url="/", status_code=303)

    @app.post("/reset", response_class=HTMLResponse)
    async def reset_state():
        """Reset all application state. Used for testing."""
        state = get_state()
        state.entity_tree = []
        state.nodes_by_id = {}
        state.editing_node_id = None
        state.current_nested_items = {}
        state.nested_edit_stack = []
        return HTMLResponse(content="OK")

    @app.get("/export")
    async def export_excel(_request: Request):
        """Export current entity data to Excel file.

        Returns an Excel workbook with one sheet per entity type,
        including all nested entities extracted recursively.
        """
        from openpyxl import Workbook

        state = get_state()
        facade = state.get_or_create_facade()

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Collect all entities by type (including nested ones)
        entities_by_type: dict[str, list[dict]] = {}

        def extract_nested_entities(data: dict, entity_type: str) -> None:
            """Recursively extract nested entities from data."""
            helper = getattr(facade, entity_type, None)
            if not helper:
                return

            for field_name, nested_type in helper.nested_fields.items():
                if field_name in data and data[field_name]:
                    nested_items = data[field_name]
                    if isinstance(nested_items, list):
                        for item in nested_items:
                            if hasattr(item, "model_dump"):
                                item_data = item.model_dump(exclude_none=True)
                            elif isinstance(item, dict):
                                item_data = item.copy()
                            else:
                                continue

                            if nested_type not in entities_by_type:
                                entities_by_type[nested_type] = []
                            entities_by_type[nested_type].append(item_data)

                            # Recursively extract from this nested item
                            extract_nested_entities(item_data, nested_type)

        # Process all root nodes
        for node in state.nodes_by_id.values():
            entity_type = node.entity_type
            if entity_type not in entities_by_type:
                entities_by_type[entity_type] = []

            if hasattr(node.instance, "model_dump"):
                data = node.instance.model_dump(exclude_none=True)
            else:
                data = {}

            entities_by_type[entity_type].append(data)

            # Extract nested entities recursively
            extract_nested_entities(data, entity_type)

        if not entities_by_type:
            # Create empty workbook with info sheet
            ws = wb.create_sheet("Info")
            ws.append(["No entities to export"])
        else:
            for entity_type, entities in entities_by_type.items():
                if not entities:
                    continue

                ws = wb.create_sheet(entity_type)

                # Get field names from helper (exclude nested fields for cleaner output)
                helper = getattr(facade, entity_type, None)
                if helper:
                    nested_fields = set(helper.nested_fields.keys())
                    columns = [f for f in helper.all_fields if f not in nested_fields]
                else:
                    # Fallback: get keys from first entity, excluding nested
                    columns = [k for k in entities[0] if not isinstance(entities[0].get(k), list)]

                # Write header
                ws.append(columns)

                # Write data rows
                for entity_data in entities:
                    row = []
                    for col in columns:
                        value = entity_data.get(col, "")
                        # Convert simple lists to comma-separated strings
                        if isinstance(value, list):
                            # Only stringify if it's a list of primitives
                            if value and not isinstance(value[0], dict):
                                value = ", ".join(str(v) for v in value)
                            else:
                                value = f"[{len(value)} items]"
                        elif isinstance(value, dict):
                            value = "[object]"
                        row.append(value)
                    ws.append(row)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"miappe_export_{state.profile}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.post("/validate", response_class=HTMLResponse)
    async def validate_form(request: Request):
        """Validate form data against MIAPPE spec.

        Returns HTML fragment with validation results.
        """
        state = get_state()
        form_data = await request.form()
        entity_type = form_data.get("_entity_type")

        if not entity_type:
            return templates.TemplateResponse(
                request,
                "components/validation_result.html",
                {
                    "valid": False,
                    "errors": [
                        {
                            "field": "_entity_type",
                            "message": "Entity type is required",
                            "rule": "required",
                        }
                    ],
                },
            )

        # Get field helper for collecting form values
        facade = ProfileFacade(profile=state.profile)
        helper = getattr(facade, entity_type)

        # Collect form values
        values = _collect_form_values(dict(form_data), helper)

        # Add nested items if in nested context or from current state
        for field_name, items in state.current_nested_items.items():
            if items:
                values[field_name] = items

        # Run validation
        errors = validate_data(values, entity_type, version="1.1")

        error_list = [{"field": e.field, "message": e.message, "rule": e.rule} for e in errors]

        return templates.TemplateResponse(
            request,
            "components/validation_result.html",
            {"valid": len(errors) == 0, "errors": error_list},
        )

    return app


def _get_field_data(helper) -> list[dict]:
    """Get field data for template rendering."""
    fields = []
    for field_name in helper.all_fields:
        info = helper.field_info(field_name)
        fields.append(
            {
                "name": field_name,
                "type": info["type"],
                "required": info["required"],
                "description": info.get("description", ""),
                "items": info.get("items"),
            }
        )
    return fields


def _is_nested_field(field: dict) -> bool:
    """Check if a field represents a nested entity (list of entities or single entity)."""
    if field["type"] == "entity":
        return True
    if field["type"] == "list":
        items = field.get("items")
        if items and items not in ("string", "int", "float", "bool"):
            return True
    return False


def _collect_form_values(form_data, helper) -> dict:
    """Collect form values into a dictionary."""
    values = {}
    for field_name in helper.all_fields:
        value = form_data.get(field_name)
        if value is None or value == "":
            continue

        info = helper.field_info(field_name)
        field_type = info["type"]

        if field_type == "integer":
            try:
                value = int(value)
            except ValueError:
                continue
        elif field_type == "float":
            try:
                value = float(value)
            except ValueError:
                continue
        elif field_type == "boolean":
            value = value.lower() in ("true", "1", "yes", "on")
        elif field_type == "list" and info.get("items") == "string":
            value = [line.strip() for line in str(value).split("\n") if line.strip()]

        values[field_name] = value

    return values


def _format_validation_errors(e: ValidationError) -> str:
    """Format validation errors for display with user-friendly messages."""
    friendly_messages = []
    for err in e.errors():
        field = ".".join(str(loc) for loc in err["loc"])
        msg = err["msg"]

        # Make common error messages more user-friendly
        if "pattern" in msg.lower() and "email" in field.lower():
            msg = "Invalid email format"
        elif "pattern" in msg.lower() and ("date" in field.lower() or "date" in msg.lower()):
            msg = "Invalid date format (use YYYY-MM-DD)"
        elif "pattern" in msg.lower() and "orcid" in field.lower():
            msg = "Invalid ORCID format (use XXXX-XXXX-XXXX-XXXX)"
        elif "pattern" in msg.lower():
            msg = "Invalid format"
        elif "required" in msg.lower():
            msg = "This field is required"

        friendly_messages.append(f"{field}: {msg}")

    return "; ".join(friendly_messages)


def _error_response(request: Request, templates: Jinja2Templates, message: str) -> HTMLResponse:
    """Create an error response with notification."""
    return templates.TemplateResponse(
        request,
        "components/notification.html",
        {
            "type": "error",
            "message": message,
        },
    )


def _get_table_columns(facade: ProfileFacade, entity_type: str) -> list[str]:
    """Get table columns for a nested entity type."""
    helper = getattr(facade, entity_type, None)
    if not helper:
        return ["value"]

    cols = list(helper.required_fields) + list(helper.optional_fields)
    nested = set(helper.nested_fields.keys())
    return [c for c in cols if c not in nested]


def _format_table_rows(items: list) -> list[dict]:
    """Format items for table row rendering."""
    rows = []
    for i, item in enumerate(items):
        if hasattr(item, "model_dump"):
            row = item.model_dump(exclude_none=True)
        elif isinstance(item, dict):
            row = item.copy()
        else:
            row = {"value": str(item)}
        row["_idx"] = i
        rows.append(row)
    return rows


def _build_breadcrumb(state: AppState) -> list[dict]:
    """Build breadcrumb navigation from nested edit stack."""
    breadcrumb = []

    # Root entity (if editing)
    if state.editing_node_id:
        node = state.nodes_by_id.get(state.editing_node_id)
        if node:
            breadcrumb.append(
                {
                    "label": node.label or node.entity_type,
                    "entity_type": node.entity_type,
                    "url": f"/form/{node.entity_type}/{node.id}",
                }
            )

    # Nested contexts
    for ctx in state.nested_edit_stack:
        breadcrumb.append(
            {
                "label": ctx.field_name,
                "entity_type": ctx.entity_type,
                "url": None,  # Table views don't have direct URL
            }
        )
        breadcrumb.append(
            {
                "label": f"{ctx.entity_type}[{ctx.row_idx}]",
                "entity_type": ctx.entity_type,
                "url": None,  # Current position
            }
        )

    return breadcrumb


app = create_app()


def run_ui(host: str = "127.0.0.1", port: int = 8080) -> None:
    """Run the MIAPPE-API web interface."""
    import uvicorn

    uvicorn.run(app, host=host, port=port)
